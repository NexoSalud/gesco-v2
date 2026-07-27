"""Standalone import script for Apoyo Administrativo from Excel.

Usage: python scripts/import_apoyo_excel.py /path/to/file.xlsx

Connects to the same DB configured in environment (DATABASE_URL).
Works standalone — no app module dependency.
"""

import asyncio
import os
import sys
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("import_apoyo_excel")

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql+asyncpg://gesco:gesco123@db:5432/gesco_v2",
)


async def run():
    import openpyxl

    from sqlalchemy import text
    from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
    from sqlalchemy.orm import sessionmaker

    filepath = sys.argv[1] if len(sys.argv) > 1 else "/tmp/import.xlsx"

    engine = create_async_engine(DATABASE_URL, echo=False)

    # Verify connection
    async with engine.begin() as conn:
        result = await conn.execute(text("SELECT NOW()"))
        print(f"✅ Conectado a DB: {result.scalar()}")

    # Read Excel
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [str(c.value).strip().upper() for c in ws[1]]
    col_map = {}
    for i, h in enumerate(headers):
        col_map[h] = i

    groups = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        perfil = str(row[col_map["PERFIL"]].value or "").strip()
        nombre = str(row[col_map["NOMBRE_COMPLETO"]].value or "").strip()
        orden_val = row[col_map["ORDEN"]].value
        actividad = str(row[col_map["ACTIVIDAD"]].value or "").strip()
        if not perfil or not nombre or not actividad:
            continue
        try:
            orden = int(orden_val) if orden_val else 0
        except (ValueError, TypeError):
            orden = 0
        key = (perfil.upper(), nombre.upper())
        if key not in groups:
            groups[key] = {"perfil": perfil, "nombre": nombre, "actividades": []}
        groups[key]["actividades"].append((orden, actividad))

    print(f"\n📊 Perfiles encontrados: {len(groups)}")

    Session = sessionmaker(bind=engine, class_=AsyncSession)
    creados = 0
    actualizados = 0
    total_actividades = 0

    async with Session() as db:
        for key, group in groups.items():
            perfil = group["perfil"]
            nombre = group["nombre"]
            actividades = sorted(group["actividades"], key=lambda x: x[0])

            # Buscar apoyo existente por nombre (case-insensitive)
            result = await db.execute(
                text("SELECT id, perfil FROM apoyo_administrativo WHERE UPPER(nombre) = :nombre"),
                {"nombre": nombre.upper()},
            )
            row = result.one_or_none()

            if row:
                apoyo_id = row[0]
                # Actualizar perfil si cambió
                if perfil and row[1] != perfil:
                    await db.execute(
                        text("UPDATE apoyo_administrativo SET perfil = :perfil WHERE id = :id"),
                        {"perfil": perfil, "id": apoyo_id},
                    )
                actualizados += 1
                action = "ACTUALIZADO"
            else:
                # Crear nuevo
                ident = f"APOYO-{nombre.split()[0][:10].upper()}-{len(nombre):03d}"
                result = await db.execute(
                    text("""
                        INSERT INTO apoyo_administrativo (nombre, identificacion, perfil, activo)
                        VALUES (:nombre, :ident, :perfil, TRUE)
                        RETURNING id
                    """),
                    {"nombre": nombre, "ident": ident, "perfil": perfil},
                )
                apoyo_id = result.scalar()
                creados += 1
                action = "CREADO"

            # Eliminar actividades existentes
            await db.execute(
                text("DELETE FROM actividades_apoyo WHERE apoyo_id = :id"),
                {"id": apoyo_id},
            )

            # Crear nuevas actividades
            for orden, descripcion in actividades:
                await db.execute(
                    text("""
                        INSERT INTO actividades_apoyo (apoyo_id, descripcion, tipo, orden)
                        VALUES (:apoyo_id, :descripcion, 'GENERAL', :orden)
                    """),
                    {"apoyo_id": apoyo_id, "descripcion": descripcion, "orden": orden},
                )
                total_actividades += 1

            print(f"  • {action}: {nombre} ({perfil}) → {len(actividades)} actividades")

        await db.commit()

    print(f"\n✅ Importación completada:")
    print(f"   Creados: {creados}")
    print(f"   Actualizados: {actualizados}")
    print(f"   Total actividades: {total_actividades}")
    print(f"   Perfiles cargados: {len(groups)}")

    await engine.dispose()


if __name__ == "__main__":
    asyncio.run(run())
