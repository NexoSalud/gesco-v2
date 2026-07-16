"""Router para importación masiva de contratos desde Excel."""

import io
import logging
import re
from datetime import datetime, date

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.contrato import Contrato
from app.models.contratista import Contratista
from app.models.pago import Pago
from app.models.planilla import Planilla
from app.models.perfil import Perfil, ActividadPerfil
from app.models.actividad_contrato import ActividadContrato


# Mapa de normalización de perfiles (variantes → nombre oficial)
PERFIL_NORMALIZATION = {
    "MEDICO": "MEDICINA",
    "MEDICINA": "MEDICINA",
    "MEDICO GENERAL": "MEDICINA",
    "MEDICO RURAL": "MEDICINA",
    "MEDICINA GENERAL": "MEDICINA",
    "ENFERMERO": "ENFERMERIA",
    "ENFERMERA": "ENFERMERIA",
    "ENFERMERIA": "ENFERMERIA",
    "ENFERMERO(A)": "ENFERMERIA",
    "ENF": "ENFERMERIA",
    "PSICOLOGO": "PSICOLOGIA",
    "PSICÓLOGO": "PSICOLOGIA",
    "PSICOLOGIA": "PSICOLOGIA",
    "ODONTOLOGO": "SALUD ORAL",
    "ODONTÓLOGO": "SALUD ORAL",
    "SALUD ORAL": "SALUD ORAL",
    "AUXILIAR ENFERMERIA": "AUXILIAR ENFERMERIA",
    "AUXILIAR DE ENFERMERÍA": "AUXILIAR ENFERMERIA",
    "AUX ENFERMERIA": "AUXILIAR ENFERMERIA",
    "AUXILIAR DE ENFERMERIA": "AUXILIAR ENFERMERIA",
    "AUXILIAR VACUNACION": "AUXILIAR VACUNACION",
    "AUX VACUNACION": "AUXILIAR VACUNACION",
    "VACUNADOR": "AUXILIAR VACUNACION",
    "GESTOR COMUNITARIO": "GESTOR COMUNITARIO",
    "SINDICATO": "SINDICATO",
    "SINDICATO_103": "SINDICATO",
    "SINDICATO_106": "SINDICATO",
    "APOYO ADMINISTRATIVO": "SINDICATO",
    "CONDUCTOR": "OTRO",
    "BACTERIÓLOGO": "OTRO",
    "BACTERIOLOGO": "OTRO",
    "TÉCNICO AMBIENTAL": "OTRO",
    "TECNICO AMBIENTAL": "OTRO",
    "TECNÓLOGO EN SISTEMAS": "OTRO",
    "TECNOLOGO SISTEMAS": "OTRO",
    "HIGIENISTA": "AUXILIAR ENFERMERIA",
    "AUXILIAR ENFERMERIA HIGIENISTA": "AUXILIAR ENFERMERIA",
    "CONDUCTOR": "TRANSPORTE",
    "TRANSPORTE": "TRANSPORTE",
}
from app.schemas.import_schema import ImportResult
from app.services.numero_letras import numero_a_letras

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/import", tags=["Importación"])

COLUMN_MAP = {
    # Legacy format (template antiguo)
    "NO. CONTRATO": "numero_contrato",
    "CONTRATISTA": "beneficiario_nombre",
    "CEDULA DE CONTRATISTA": "cedula_contratista",
    "LUGAR DE EXPEDICIÓN": "lugar_expedicion",
    "TELEFONO": "telefono",
    "DIRECCION": "direccion",
    "CORREO": "correo",
    "TÍTULO": "perfil",
    "VALOR DEL CONTRATO": "monto_total",
    "CUOTAS": "cuotas",
    "VIGENCIA DEL CONTRATO": "vigencia_texto",
    "OBJETO DEL CONTRATO": "objeto",
    "SUPERVISOR": "supervisor",
    "CEDULA SUPERVISOR": "cedula_supervisor",
    "No. CDP": "no_cdp",

    # Plantilla de supervisiones (nuevo formato)
    "N° DE CONTRATO": "numero_contrato",
    "NOMBRE CONTRATISTA": "beneficiario_nombre",
    "No. DE IDENTIFICACIÓN": "cedula_contratista",
    "EXPEDIDA EN": "lugar_expedicion",
    "No. TELÉFONO y/o CELULAR": "telefono",
    "DIRECCION": "direccion",
    "PERFIL": "perfil",
    "VALOR TOTAL DEL CONTRATO": "monto_total",
    "VALOR FINAL DEL CONTRATO": "monto_total",
    "OBJETO DEL CONTRATO": "objeto",
    "SUPERVISOR": "supervisor",
    "CDP  No.": "no_cdp",
    "CRP No.": "rp",
    "IMPUTACIÓN PRESUPUESTAL": "rubro",
    "UNIDAD DE ATENCION": "unidad_atencion",
    "FECHA DE INICIO DEL CONTRATO": "fecha_inicio",
    "FECHA TERMINACION DEL CONTRATO": "fecha_fin",
    "ESTADO": "estado_contrato",
    "TIPO DE PERSONA": "tipo_persona",
    "RESOLUCION": "resolucion_codigo",

    # Columnas de pagos
    "TIPO DE INFORME": "tipo_informe",
    "PERIODO INFORME DESDE": "periodo_desde",
    "PERIODO INFORME HASTA": "periodo_hasta",
    "VALOR A PAGAR": "valor_a_pagar",
    "N° FOLIOS": "folios",
    "OBSERVACIONES": "observaciones",
    "ACTIVIDADES": "actividades",
    "PAGO No": "pago_numero",

    # Nuevas columnas de pago (supervisión)
    "FECHA FIRMA": "fecha_firma",
    "OTRO SI": "otro_si",
    "VALOR PAGADO HISTORICO": "valor_pagado_historico",
    "ANEXA CERTIFICACION": "anexa_cert",

    # Planilla de seguridad social
    "PLANILLA No": "planilla_no",
    "PERIODO COTIZADO": "periodo_cotizado",
    "IBC": "ibc",
    "EPS NOMBRE": "eps_nombre",
    "EPS VALOR": "eps_valor",
    "ARL NOMBRE": "arl_nombre",
    "ARL VALOR": "arl_valor",
    "AFP NOMBRE": "afp_nombre",
    "AFP VALOR": "afp_valor",
    "CCF NOMBRE": "ccf_nombre",
    "CCF VALOR": "ccf_valor",
    "SENA VALOR": "sena_valor",
    "ICBF VALOR": "icbf_valor",

    # Nuevas columnas de contrato (supervisión PDF)
    "CODIGO CIIU": "codigo_ciiu",
    "NIVEL SUPERVISOR": "nivel_prof_supervisor",
    "INTERVENTOR": "interventor",
    "NIVEL INTERVENTOR": "nivel_prof_interventor",
    "TIEMPO ADICION": "tiempo_adicion",
    "VALOR FINAL": "valor_final",
    "FORMA PAGO": "forma_pago",

    # UNSPSC
    "CODIGO UNSPSC": "codigo_unspsc",
    "DESCRIPCION UNSPSC": "descripcion_unspsc",
}

_TITLE_NORMALIZE_RE = re.compile(r"[^A-Z0-9ÁÉÍÓÚÑ ]")


def _normalize_title(title: str) -> str:
    t = title.strip().upper()
    t = _TITLE_NORMALIZE_RE.sub("", t)
    return re.sub(r"\s+", " ", t)


def _find_column_index(headers: list[str], target: str) -> int | None:
    target_norm = _normalize_title(target)
    for i, h in enumerate(headers):
        if _normalize_title(str(h)) == target_norm:
            return i
    return None


def _parse_number(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("$", "").replace(" ", "").replace("\xa0", "")
    if "," in s and "." in s:
        if s.rfind(".") > s.rfind(","):
            s = s.replace(",", "")
        else:
            s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "." in s:
        parts = s.split(".")
        if len(parts) > 2:
            s = s.replace(".", "")
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def _clean_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _get_col(col_indices: dict[str, int], row: tuple, *names: str):
    """Read a cell value trying multiple column names."""
    for name in names:
        idx = col_indices.get(name)
        if idx is not None and idx < len(row):
            val = row[idx]
            return val
    return None


def _parse_date(val) -> date | None:
    """Parse a date from various formats."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val if isinstance(val, date) else val.date()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _map_estado(val) -> str:
    """Map Excel estado to internal estado."""
    if val is None:
        return "EN_PROCESO"
    s = str(val).strip().upper()
    if s in ("ACTIVO",):
        return "ACTIVO"
    if s in ("FINALIZADO",):
        return "FINALIZADO"
    if s in ("ANULADO",):
        return "ANULADO"
    return "EN_PROCESO"


@router.post("/excel", response_model=ImportResult)
async def importar_contratos_excel(
    resolucion_id: int = Query(..., description="ID de la resolución destino"),
    file: UploadFile = File(..., description="Archivo Excel (.xlsx)"),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "El archivo debe ser .xlsx")

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(400, "El Excel no tiene hojas activas")
    except Exception as e:
        raise HTTPException(400, f"Error al leer el Excel: {e}")

    rows_iter = ws.iter_rows(values_only=True)

    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        raise HTTPException(400, "El archivo Excel está vacío")

    headers = [str(h) if h else "" for h in raw_headers]

    col_indices: dict[str, int] = {}
    for col_name in COLUMN_MAP:
        idx = _find_column_index(headers, col_name)
        if idx is not None:
            col_indices[col_name] = idx

    if not col_indices:
        raise HTTPException(
            400,
            "No se encontraron columnas reconocidas en el Excel. "
            "Verifique que los encabezados coincidan con el formato esperado.",
        )

    result_db = await db.execute(select(Perfil))
    perfiles_existentes = {p.nombre.upper(): p.nombre for p in result_db.scalars().all()}

    result_summary = ImportResult()
    contratos_creados: set[str] = set()  # track created contracts for dedup

    for fila_idx, row in enumerate(rows_iter, start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        try:
            # ─── Número de contrato ──
            numero_contrato = _clean_str(_get_col(col_indices, row, "N° DE CONTRATO", "NO. CONTRATO"))
            if not numero_contrato:
                continue

            result_summary.total += 1
            es_nuevo_contrato = numero_contrato not in contratos_creados

            # ─── Si es nuevo contrato, crear contratista y contrato ──
            if es_nuevo_contrato:
                beneficiario_nombre = _clean_str(_get_col(col_indices, row, "NOMBRE CONTRATISTA", "CONTRATISTA"))
                cedula_contratista = _clean_str(_get_col(col_indices, row, "No. DE IDENTIFICACIÓN", "CEDULA DE CONTRATISTA"))
                lugar_expedicion = _clean_str(_get_col(col_indices, row, "EXPEDIDA EN", "LUGAR DE EXPEDICIÓN"))
                telefono = _clean_str(_get_col(col_indices, row, "No. TELÉFONO y/o CELULAR", "TELEFONO"))
                direccion = _clean_str(_get_col(col_indices, row, "DIRECCION"))
                correo = _clean_str(_get_col(col_indices, row, "CORREO"))
                tipo_persona = _clean_str(_get_col(col_indices, row, "TIPO DE PERSONA"))

                if not cedula_contratista:
                    result_summary.errors.append({"fila": fila_idx, "numero_contrato": numero_contrato, "error": "Cédula del contratista vacía"})
                    result_summary.skipped += 1
                    continue

                # ─── Perfil ──
                perfil_raw = _clean_str(_get_col(col_indices, row, "PERFIL", "TÍTULO"))
                perfil_normalized = None
                if perfil_raw:
                    perfil_upper = perfil_raw.upper().strip()
                    # 1. Intentar normalización exacta por mapa
                    if perfil_upper in PERFIL_NORMALIZATION:
                        perfil_normalized = PERFIL_NORMALIZATION[perfil_upper]
                    # 2. Buscar en perfiles existentes en BD
                    elif perfil_upper in perfiles_existentes:
                        perfil_normalized = perfiles_existentes[perfil_upper]
                    else:
                        # 3. Búsqueda parcial contra nombres existentes
                        for p_name in perfiles_existentes.values():
                            if perfil_upper in p_name.upper() or p_name.upper() in perfil_upper:
                                perfil_normalized = p_name
                                break
                        # 4. Último recurso: usar raw
                        if not perfil_normalized:
                            perfil_normalized = perfil_raw.upper()

                monto_total = _parse_number(_get_col(col_indices, row, "VALOR FINAL DEL CONTRATO", "VALOR TOTAL DEL CONTRATO", "VALOR DEL CONTRATO"))
                fecha_inicio = _parse_date(_get_col(col_indices, row, "FECHA DE INICIO DEL CONTRATO"))
                fecha_fin = _parse_date(_get_col(col_indices, row, "FECHA TERMINACION DEL CONTRATO"))
                objeto = _clean_str(_get_col(col_indices, row, "OBJETO DEL CONTRATO"))
                supervisor = _clean_str(_get_col(col_indices, row, "SUPERVISOR"))
                cedula_supervisor = _clean_str(_get_col(col_indices, row, "CEDULA SUPERVISOR"))
                no_cdp = _clean_str(_get_col(col_indices, row, "CDP  No.", "No. CDP"))
                rp = _clean_str(_get_col(col_indices, row, "CRP No."))
                rubro = _clean_str(_get_col(col_indices, row, "IMPUTACIÓN PRESUPUESTAL"))
                unidad_atencion = _clean_str(_get_col(col_indices, row, "UNIDAD DE ATENCION"))

                # Nuevos campos PDF supervisión
                codigo_ciiu = _clean_str(_get_col(col_indices, row, "CODIGO CIIU")) or ""
                nivel_prof_supervisor = _clean_str(_get_col(col_indices, row, "NIVEL SUPERVISOR")) or ""
                interventor = _clean_str(_get_col(col_indices, row, "INTERVENTOR")) or ""
                nivel_prof_interventor = _clean_str(_get_col(col_indices, row, "NIVEL INTERVENTOR")) or ""
                tiempo_adicion = _clean_str(_get_col(col_indices, row, "TIEMPO ADICION")) or ""
                valor_final = _parse_number(_get_col(col_indices, row, "VALOR FINAL")) or 0
                forma_pago = _clean_str(_get_col(col_indices, row, "FORMA PAGO")) or ""
                codigo_unspsc_import = _clean_str(_get_col(col_indices, row, "CODIGO UNSPSC")) or ""
                descripcion_unspsc_import = _clean_str(_get_col(col_indices, row, "DESCRIPCION UNSPSC")) or ""

                # Cuotas: ya no se usan, se deja en 0
                # El progreso se mide por (total pagado / monto_total)
                cuotas_txt = None
                cuotas_total = 0

                estado_raw = _get_col(col_indices, row, "ESTADO", "estado_contrato")
                estado = _map_estado(estado_raw)

                # ─── Contratista ──
                result_db2 = await db.execute(select(Contratista).where(Contratista.identificacion == cedula_contratista))
                contratista_obj = result_db2.scalar_one_or_none()

                if contratista_obj:
                    contratista_id = contratista_obj.id
                    if beneficiario_nombre:
                        contratista_obj.nombre = beneficiario_nombre.upper().strip('"')
                    if lugar_expedicion:
                        contratista_obj.expedida_en = lugar_expedicion.upper()
                    if telefono:
                        contratista_obj.telefono = telefono
                    if direccion:
                        contratista_obj.direccion = direccion.upper()
                    if correo:
                        contratista_obj.correo = correo.lower()
                    if tipo_persona:
                        contratista_obj.tipo_persona = tipo_persona.upper()
                else:
                    contratista_obj = Contratista(
                        identificacion=cedula_contratista,
                        nombre=(beneficiario_nombre or "").upper().strip('"'),
                        expedida_en=(lugar_expedicion or "").upper(),
                        telefono=telefono or "",
                        direccion=(direccion or "").upper(),
                        correo=(correo or "").lower(),
                        tipo_persona=(tipo_persona or "").upper(),
                    )
                    db.add(contratista_obj)
                    await db.flush()
                    contratista_id = contratista_obj.id

                # ─── Verificar duplicado en BD ──
                existing = await db.execute(select(Contrato).where(Contrato.numero_contrato == numero_contrato))
                if existing.scalar_one_or_none():
                    result_summary.errors.append({"fila": fila_idx, "numero_contrato": numero_contrato, "error": "El contrato ya existe en BD"})
                    result_summary.skipped += 1
                    continue

                # Heredar UNSPSC del perfil si no se especificó en Excel
                if not codigo_unspsc_import and perfil_normalized:
                    result_perf_unspsc = await db.execute(
                        select(Perfil).where(Perfil.nombre == perfil_normalized)
                    )
                    perf_obj = result_perf_unspsc.scalar_one_or_none()
                    if perf_obj:
                        if perf_obj.codigo_unspsc:
                            codigo_unspsc_import = perf_obj.codigo_unspsc
                        if perf_obj.descripcion_unspsc:
                            descripcion_unspsc_import = perf_obj.descripcion_unspsc

                # ─── Crear contrato ──
                valor_letras = numero_a_letras(monto_total)
                contrato_obj = Contrato(
                    resolucion_id=resolucion_id,
                    contratista_id=contratista_id,
                    numero_contrato=numero_contrato,
                    perfil=perfil_normalized,
                    estado=estado,
                    objeto=objeto or "",
                    monto_total=monto_total,
                    valor_letras=valor_letras,
                    supervisor=supervisor or "",
                    cedula_supervisor=cedula_supervisor or "",
                    no_cdp=no_cdp or "",
                    rp=rp or "",
                    rubro=rubro or "",
                    unidad_atencion=unidad_atencion or "",
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    codigo_ciiu=codigo_ciiu,
                    codigo_unspsc=codigo_unspsc_import or None,
                    descripcion_unspsc=descripcion_unspsc_import or None,
                    nivel_prof_supervisor=nivel_prof_supervisor,
                    interventor=interventor,
                    nivel_prof_interventor=nivel_prof_interventor,
                    tiempo_adicion=tiempo_adicion,
                    valor_final=valor_final,
                    forma_pago=forma_pago,
                )
                db.add(contrato_obj)
                await db.flush()
                contratos_creados.add(numero_contrato)
                result_summary.created += 1

                # Heredar actividades del perfil al contrato
                if perfil_normalized:
                    result_p = await db.execute(
                        select(Perfil).where(Perfil.nombre == perfil_normalized)
                    )
                    perfil_obj = result_p.scalar_one_or_none()
                    if perfil_obj:
                        result_acts = await db.execute(
                            select(ActividadPerfil).where(ActividadPerfil.perfil_id == perfil_obj.id).order_by(ActividadPerfil.orden)
                        )
                        for ap in result_acts.scalars().all():
                            ac = ActividadContrato(
                                contrato_id=numero_contrato,
                                descripcion=ap.descripcion,
                                tipo="GENERAL",
                                orden=ap.orden,
                            )
                            db.add(ac)
                        await db.flush()
            else:
                # Contrato ya existe en este batch — obtenerlo
                existing = await db.execute(select(Contrato).where(Contrato.numero_contrato == numero_contrato))
                contrato_obj = existing.scalar_one_or_none()
                if not contrato_obj:
                    result_summary.errors.append({"fila": fila_idx, "numero_contrato": numero_contrato, "error": "Contrato no encontrado para crear pago"})
                    result_summary.skipped += 1
                    continue

            # ─── DATOS DEL PAGO ──
            valor_a_pagar = _parse_number(_get_col(col_indices, row, "VALOR A PAGAR"))

            # Saltar creación de pago si el valor es 0 o negativo
            if valor_a_pagar > 0:
                tipo_informe = _clean_str(_get_col(col_indices, row, "TIPO DE INFORME")) or "SUPERVISION"
                periodo_desde = _parse_date(_get_col(col_indices, row, "PERIODO INFORME DESDE"))
                periodo_hasta = _parse_date(_get_col(col_indices, row, "PERIODO INFORME HASTA"))
                folios = _clean_str(_get_col(col_indices, row, "N° FOLIOS"))
                observaciones = _clean_str(_get_col(col_indices, row, "OBSERVACIONES"))
                actividades = _clean_str(_get_col(col_indices, row, "ACTIVIDADES"))
                fecha_firma = _parse_date(_get_col(col_indices, row, "FECHA FIRMA"))
                otro_si = _clean_str(_get_col(col_indices, row, "OTRO SI"))
                valor_pagado_historico = _parse_number(_get_col(col_indices, row, "VALOR PAGADO HISTORICO"))
                anexa_cert = _clean_str(_get_col(col_indices, row, "ANEXA CERTIFICACION"))

                # Fecha dummy si vacía (para permitir registro)
                if periodo_desde is None:
                    periodo_desde = date(2026, 1, 1)
                if periodo_hasta is None:
                    periodo_hasta = date(2026, 1, 31)

                # Número de pago: forzar conversión a int
                pago_numero_raw = _get_col(col_indices, row, "PAGO No")
                numero_pago = 1
                if pago_numero_raw is not None:
                    try:
                        numero_pago = int(float(str(pago_numero_raw)))
                    except (ValueError, TypeError):
                        pass

                # Crear pago
                pago = Pago(
                    contrato_id=numero_contrato,
                    numero_pago=numero_pago,
                    tipo_informe=tipo_informe,
                    periodo_desde=periodo_desde,
                    periodo_hasta=periodo_hasta,
                    fecha_firma=fecha_firma,
                    valor_a_pagar=valor_a_pagar,
                    otro_si=otro_si,
                    valor_pagado=valor_pagado_historico if valor_pagado_historico else None,
                    anexa_cert=anexa_cert,
                    folios=folios or "",
                    observaciones=observaciones or "",
                    actividades=actividades or "",
                )
                db.add(pago)
                await db.flush()

                # ─── PLANILLA DE SEGURIDAD SOCIAL ──
                planilla_no = _clean_str(_get_col(col_indices, row, "PLANILLA No"))
                if planilla_no:  # Solo crear planilla si hay No. de planilla
                    periodo_cotizado = _clean_str(_get_col(col_indices, row, "PERIODO COTIZADO")) or ""
                    ibc = str(_parse_number(_get_col(col_indices, row, "IBC")) or 0)
                    eps_nombre = _clean_str(_get_col(col_indices, row, "EPS NOMBRE")) or ""
                    eps_valor = _parse_number(_get_col(col_indices, row, "EPS VALOR"))
                    arl_nombre = _clean_str(_get_col(col_indices, row, "ARL NOMBRE")) or ""
                    arl_valor = _parse_number(_get_col(col_indices, row, "ARL VALOR"))
                    afp_nombre = _clean_str(_get_col(col_indices, row, "AFP NOMBRE")) or ""
                    afp_valor = _parse_number(_get_col(col_indices, row, "AFP VALOR"))
                    ccf_nombre = _clean_str(_get_col(col_indices, row, "CCF NOMBRE")) or ""
                    ccf_valor = _parse_number(_get_col(col_indices, row, "CCF VALOR"))
                    sena_valor = _parse_number(_get_col(col_indices, row, "SENA VALOR"))
                    icbf_valor = _parse_number(_get_col(col_indices, row, "ICBF VALOR"))
                    valor_total_planilla = eps_valor + arl_valor + afp_valor + ccf_valor + sena_valor + icbf_valor

                    planilla = Planilla(
                        pago_id=pago.id,
                        planilla_no=planilla_no or "",
                        periodo_cotizado=periodo_cotizado,
                        ibc=ibc,
                        eps_nombre=eps_nombre,
                        eps_valor=eps_valor,
                        arl_nombre=arl_nombre,
                        arl_valor=arl_valor,
                        afp_nombre=afp_nombre,
                        afp_valor=afp_valor,
                        ccf_nombre=ccf_nombre,
                        ccf_valor=ccf_valor,
                        sena_valor=sena_valor,
                        icbf_valor=icbf_valor,
                        valor_total=valor_total_planilla,
                    )
                    db.add(planilla)

        except Exception as e:
            logger.exception(f"Error procesando fila {fila_idx}")
            err_numero = ""
            for col_name in ("N° DE CONTRATO", "NO. CONTRATO"):
                idx = col_indices.get(col_name)
                if idx is not None:
                    try:
                        err_numero = str(row[idx])
                        break
                    except Exception:
                        continue
            result_summary.errors.append({"fila": fila_idx, "numero_contrato": err_numero or "", "error": str(e)})
            result_summary.skipped += 1
            continue

    await db.commit()
    return result_summary



