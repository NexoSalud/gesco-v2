"""Router para exportaciones (Excel, PDFs masivos, alertas)."""

from datetime import date, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
import zipfile
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app.models.contrato import Contrato
from app.models.pago import Pago
from app.models.actividad_supervision import ActividadSupervision
from app.services.excel_service import exportar_resolucion_excel
from app.services.pdf_generator import generar_supervision_pdf

router = APIRouter(prefix="/api/v1/export", tags=["Exportación"])


@router.get("/resolucion/{resolucion_id}/excel")
async def exportar_resolucion_excel_endpoint(
    resolucion_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Exporta los contratos de una resolución a Excel."""
    result = await db.execute(
        text("""
            SELECT c.*, co.nombre as beneficiario, co.identificacion as cedula_contratista
            FROM contratos c
            LEFT JOIN contratistas co ON co.id = c.contratista_id
            WHERE c.resolucion_id = :rid
            ORDER BY c.numero_contrato
        """),
        {"rid": resolucion_id},
    )
    rows = [dict(r._mapping) for r in result.fetchall()]

    excel_bytes = exportar_resolucion_excel(rows)

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=contratos_resolucion_{resolucion_id}.xlsx"},
    )


@router.get("/resolucion/{resolucion_id}/pdfs-masivos")
async def generar_pdfs_masivos(
    resolucion_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Genera un ZIP con los PDFs de supervisión de todos los pagos de una resolución."""
    result = await db.execute(
        select(Contrato)
        .options(
            selectinload(Contrato.contratista_rel),
            selectinload(Contrato.pagos).selectinload(Pago.planillas),
        )
        .where(Contrato.resolucion_id == resolucion_id, Contrato.estado != "ANULADO")
    )
    contratos = result.scalars().all()

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for contrato in contratos:
            contratista = contrato.contratista_rel
            for pago in (contrato.pagos or []):
                # Calcular valor pagado histórico y saldo
                pagos_previos = [pg for pg in (contrato.pagos or []) if pg.id != pago.id]
                valor_pagado_historico = sum(float(pg.valor_a_pagar or 0) for pg in pagos_previos)
                valor_final_pdf = float(contrato.valor_final or contrato.monto_total or 0)
                causado_hasta_hoy = valor_pagado_historico + float(pago.valor_a_pagar or 0)
                saldo_a_pagar = max(0, valor_final_pdf - causado_hasta_hoy)

                data_contrato = {
                    "numero_contrato": contrato.numero_contrato,
                    "perfil": contrato.perfil or "",
                    "nombre_contratista": contratista.nombre if contratista else "",
                    "identificacion": contratista.identificacion if contratista else "",
                    "lugar_expedicion": contratista.expedida_en if contratista else "",
                    "telefono": contratista.telefono if contratista else "",
                    "direccion": contratista.direccion if contratista else "",
                    "tipo_persona": contratista.tipo_persona if contratista else "",
                    "supervisor": contrato.supervisor or "",
                    "cedula_supervisor": contrato.cedula_supervisor or "",
                    "cargo_supervisor": contrato.cargo_supervisor or "",
                    "unidad_atencion": contrato.unidad_atencion or "",
                    "no_cdp": contrato.no_cdp or "",
                    "rp": contrato.rp or "",
                    "rubro": contrato.rubro or "",
                    "monto_total": contrato.monto_total,
                    "fecha_inicio": str(contrato.fecha_inicio) if contrato.fecha_inicio else "",
                    "fecha_fin": str(contrato.fecha_fin) if contrato.fecha_fin else "",
                    "objeto": contrato.objeto or "",
                    "codigo_ciiu": contrato.codigo_ciiu or "",
                    "nivel_prof_supervisor": contrato.nivel_prof_supervisor or "",
                    "interventor": contrato.interventor or "",
                    "nivel_prof_interventor": contrato.nivel_prof_interventor or "",
                    "imputacion": contrato.imputacion or "",
                    "tiempo_adicion": contrato.tiempo_adicion or "",
                    "valor_final": valor_final_pdf,
                    "forma_pago": contrato.forma_pago or "",
                }
                data_pago = {
                    "numero_pago": pago.numero_pago,
                    "tipo_informe": pago.tipo_informe or "SUPERVISION",
                    "periodo_desde": str(pago.periodo_desde) if pago.periodo_desde else "",
                    "periodo_hasta": str(pago.periodo_hasta) if pago.periodo_hasta else "",
                    "valor_a_pagar": pago.valor_a_pagar,
                    "otro_si": pago.otro_si or 0,
                    "valor_pagado_historico": valor_pagado_historico,
                    "saldo_a_pagar": saldo_a_pagar,
                    "anexa_cert": pago.anexa_cert or "",
                    "actividades": pago.actividades or "",
                    "observaciones": pago.observaciones or "",
                    "folios": pago.folios or "",
                }
                planillas_list = [
                    {
                        "planilla_no": pl.planilla_no or "",
                        "periodo_cotizado": pl.periodo_cotizado or "",
                        "ibc": pl.ibc or "0",
                        "eps_nombre": pl.eps_nombre or "",
                        "eps_valor": pl.eps_valor or 0,
                        "arl_nombre": pl.arl_nombre or "",
                        "arl_valor": pl.arl_valor or 0,
                        "afp_nombre": pl.afp_nombre or "",
                        "afp_valor": pl.afp_valor or 0,
                        "ccf_nombre": pl.ccf_nombre or "",
                        "ccf_valor": pl.ccf_valor or 0,
                    }
                    for pl in pago.planillas
                ]
                try:
                    # Load actividades de supervisión
                    acts_sup = await db.execute(
                        select(ActividadSupervision)
                        .where(ActividadSupervision.pago_id == pago.id)
                        .order_by(ActividadSupervision.orden)
                    )
                    actividades_supervision = [
                        {"descripcion": a.descripcion, "cumple": a.cumple}
                        for a in acts_sup.scalars().all()
                    ]
                    pdf_bytes = generar_supervision_pdf(data_contrato, data_pago, planillas_list,
                                                        actividades_supervision=actividades_supervision)
                    filename = f"Supervision_{contrato.numero_contrato}_Pago{pago.numero_pago}.pdf"
                    zf.writestr(filename, pdf_bytes)
                except Exception:
                    continue

    zip_buffer.seek(0)
    return Response(
        content=zip_buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=supervisiones_resolucion_{resolucion_id}.zip"},
    )


@router.get("/alertas")
async def alertas_vencimiento(
    dias: int = Query(30, description="Días para considerar próximo vencimiento"),
    db: AsyncSession = Depends(get_db),
):
    """Alertas de contratos próximos a vencerse."""
    hoy = date.today()
    limite = hoy + timedelta(days=dias)
    result = await db.execute(
        text("""
            SELECT numero_contrato, perfil, fecha_fin, estado
            FROM contratos
            WHERE estado = 'ACTIVO'
              AND fecha_fin IS NOT NULL
              AND fecha_fin BETWEEN :hoy AND :limite
            ORDER BY fecha_fin
        """),
        {"hoy": hoy, "limite": limite},
    )
    return [dict(r._mapping) for r in result.fetchall()]


@router.get("/resolucion/{resolucion_id}/analytics")
async def resolucion_analytics(
    resolucion_id: int,
    db: AsyncSession = Depends(get_db),
):
    """KPIs detallados de una resolución (analytics)."""
    hoy = date.today().isoformat()
    treinta = (date.today() + timedelta(days=30)).isoformat()

    # ── Totales ──
    total_contratos = await db.execute(
        text("""SELECT COUNT(*) FROM contratos
                 WHERE resolucion_id = :rid AND estado <> 'ANULADO'"""),
        {"rid": resolucion_id},
    )
    total_contratos = total_contratos.scalar() or 0

    contratos_activos = await db.execute(
        text("""SELECT COUNT(*) FROM contratos
                 WHERE resolucion_id = :rid AND estado <> 'ANULADO'
                   AND fecha_inicio::date <= :hoy AND fecha_fin::date >= :hoy"""),
        {"rid": resolucion_id, "hoy": hoy},
    )
    contratos_activos = contratos_activos.scalar() or 0

    contratos_por_vencer = await db.execute(
        text("""SELECT COUNT(*) FROM contratos
                 WHERE resolucion_id = :rid AND estado <> 'ANULADO'
                   AND fecha_fin::date BETWEEN :hoy AND :treinta"""),
        {"rid": resolucion_id, "hoy": hoy, "treinta": treinta},
    )
    contratos_por_vencer = contratos_por_vencer.scalar() or 0

    contratos_vencidos = await db.execute(
        text("""SELECT COUNT(*) FROM contratos
                 WHERE resolucion_id = :rid AND estado <> 'ANULADO'
                   AND fecha_fin::date < :hoy"""),
        {"rid": resolucion_id, "hoy": hoy},
    )
    contratos_vencidos = contratos_vencidos.scalar() or 0

    total_anulados = await db.execute(
        text("""SELECT COUNT(*) FROM contratos
                 WHERE resolucion_id = :rid AND estado = 'ANULADO'"""),
        {"rid": resolucion_id},
    )
    total_anulados = total_anulados.scalar() or 0

    # ── Profesionales por tipo (perfil) ──
    rows_prof = await db.execute(
        text("""
            SELECT
                COALESCE(NULLIF(TRIM(c.perfil), ''), 'SIN ESPECIFICAR') AS tipo,
                COUNT(*) AS total,
                COALESCE(SUM(c.monto_total + c.monto_transporte), 0) AS valor
            FROM contratos c
            WHERE c.resolucion_id = :rid AND c.estado <> 'ANULADO'
            GROUP BY tipo
            ORDER BY total DESC
        """),
        {"rid": resolucion_id},
    )
    profesionales_por_tipo = [
        {"tipo": r.tipo, "total": r.total, "valor": float(r.valor)}
        for r in rows_prof.fetchall()
    ]

    # ── Próximos a vencerse (hoy hasta hoy+30) ──
    rows_prox = await db.execute(
        text("""
            SELECT
                c.numero_contrato,
                COALESCE(co.nombre, 'SIN CONTRATISTA') AS beneficiario,
                c.fecha_fin::text AS fecha_fin,
                (c.fecha_fin - CURRENT_DATE) AS dias_restantes
            FROM contratos c
            LEFT JOIN contratistas co ON co.id = c.contratista_id
            WHERE c.resolucion_id = :rid AND c.estado <> 'ANULADO'
              AND c.fecha_fin::date BETWEEN :hoy AND :treinta
            ORDER BY c.fecha_fin
        """),
        {"rid": resolucion_id, "hoy": hoy, "treinta": treinta},
    )
    proximos_vencer = [
        {
            "numero_contrato": r.numero_contrato,
            "beneficiario": r.beneficiario,
            "fecha_fin": r.fecha_fin,
            "dias_restantes": r.dias_restantes if r.dias_restantes is not None else 0,
        }
        for r in rows_prox.fetchall()
    ]

    # ── Motivos de anulación ──
    rows_motivos = await db.execute(
        text("""
            SELECT
                COALESCE(NULLIF(TRIM(c.motivo_anulacion), ''), 'Sin especificar') AS motivo,
                COUNT(*) AS total
            FROM contratos c
            WHERE c.resolucion_id = :rid AND c.estado = 'ANULADO'
            GROUP BY motivo
            ORDER BY total DESC
        """),
        {"rid": resolucion_id},
    )
    motivos_anulacion = [
        {"motivo": r.motivo, "total": r.total}
        for r in rows_motivos.fetchall()
    ]

    # ── Contratos por unidad de atención ──
    rows_unidad = await db.execute(
        text("""
            SELECT
                COALESCE(NULLIF(TRIM(c.unidad_atencion), ''), 'SIN ESPECIFICAR') AS municipio,
                COUNT(*) AS total,
                COUNT(CASE WHEN c.fecha_inicio::date <= :hoy_a AND c.fecha_fin::date >= :hoy_a THEN 1 END) AS activos,
                COALESCE(SUM(CASE WHEN c.estado <> 'ANULADO' THEN c.monto_total + c.monto_transporte ELSE 0 END), 0) AS valor
            FROM contratos c
            WHERE c.resolucion_id = :rid AND c.estado <> 'ANULADO'
            GROUP BY municipio
            ORDER BY activos DESC, total DESC
        """),
        {"rid": resolucion_id, "hoy_a": hoy},
    )
    contratos_por_unidad = [
        {"municipio": r.municipio, "total": r.total, "activos": r.activos, "valor": float(r.valor)}
        for r in rows_unidad.fetchall()
    ]

    return {
        "total_contratos": total_contratos,
        "contratos_activos": contratos_activos,
        "contratos_por_vencer": contratos_por_vencer,
        "contratos_vencidos": contratos_vencidos,
        "total_anulados": total_anulados,
        "profesionales_por_tipo": profesionales_por_tipo,
        "proximos_vencer": proximos_vencer,
        "motivos_anulacion": motivos_anulacion,
        "contratos_por_unidad": contratos_por_unidad,
    }


@router.get("/dashboard-global")
async def dashboard_global(db: AsyncSession = Depends(get_db)):
    """KPIs de la resolución activa. Solo una resolución activa a la vez."""
    stats = await db.execute(text("""
        SELECT
            r.id AS resolucion_activa_id,
            r.codigo AS resolucion_activa_codigo,
            r.presupuesto AS presupuesto_global,
            r.indirect_percentage,
            COUNT(c.id) AS total_contratos,
            COALESCE(SUM(CASE WHEN c.estado <> 'ANULADO' THEN c.monto_total + c.monto_transporte ELSE 0 END), 0) AS total_comprometido,
            COUNT(CASE WHEN c.estado = 'ANULADO' THEN 1 END) AS total_anulados,
            COUNT(CASE WHEN c.estado = 'FINALIZADO' THEN 1 END) AS total_finalizados,
            COUNT(CASE WHEN c.estado = 'EN_PROCESO' THEN 1 END) AS en_proceso,
            COUNT(CASE WHEN c.estado = 'ACTIVO' THEN 1 END) AS activos,
            COALESCE(AVG(CASE WHEN c.estado <> 'ANULADO' THEN c.monto_total END), 0) AS promedio_valor,
            COALESCE(SUM(c.cuotas_total), 0) AS cuotas_totales_global,
            COALESCE(SUM(c.cuotas_pagadas), 0) AS cuotas_pagadas_global
        FROM resoluciones r
        LEFT JOIN contratos c ON c.resolucion_id = r.id
        WHERE r.activa = TRUE
        GROUP BY r.id, r.codigo, r.presupuesto, r.indirect_percentage
    """))
    s = stats.fetchone()
    if not s:
        return {
            "resolucion_activa_id": None,
            "resolucion_activa_codigo": None,
            "total_resoluciones": 0,
            "total_contratos": 0,
            "total_comprometido": 0,
            "presupuesto_global": 0,
            "saldo_global": 0,
            "total_anulados": 0,
            "total_finalizados": 0,
            "en_proceso": 0,
            "activos": 0,
            "promedio_valor": 0,
            "cuotas_totales_global": 0,
            "cuotas_pagadas_global": 0,
        }
    return {
        "resolucion_activa_id": s.resolucion_activa_id,
        "resolucion_activa_codigo": s.resolucion_activa_codigo,
        "total_resoluciones": 1,
        "total_contratos": s.total_contratos,
        "total_comprometido": float(s.total_comprometido),
        "presupuesto_global": float(s.presupuesto_global),
        "saldo_global": float(s.presupuesto_global - s.total_comprometido),
        "total_anulados": s.total_anulados,
        "total_finalizados": s.total_finalizados,
        "en_proceso": s.en_proceso,
        "activos": s.activos,
        "promedio_valor": float(s.promedio_valor),
        "cuotas_totales_global": s.cuotas_totales_global,
        "cuotas_pagadas_global": s.cuotas_pagadas_global,
        "indirect_percentage": float(s.indirect_percentage),
    }


@router.get("/plantilla-importacion")
async def descargar_plantilla_importacion():
    """Descarga plantilla Excel para importación masiva de contratos con supervisiones."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Importación"

    encabezados = [
        "N° DE CONTRATO",
        "NOMBRE CONTRATISTA",
        "No. DE IDENTIFICACIÓN",
        "EXPEDIDA EN",
        "No. TELÉFONO y/o CELULAR",
        "DIRECCION",
        "TIPO DE PERSONA",
        "PERFIL",
        "VALOR TOTAL DEL CONTRATO",
        "VALOR FINAL",
        "OBJETO DEL CONTRATO",
        "SUPERVISOR",
        "CEDULA SUPERVISOR",
        "NIVEL SUPERVISOR",
        "INTERVENTOR",
        "NIVEL INTERVENTOR",
        "CDP No.",
        "CRP No.",
        "IMPUTACIÓN PRESUPUESTAL",
        "UNIDAD DE ATENCION",
        "CODIGO CIIU",
        "TIEMPO ADICION",
        "FORMA PAGO",
        "CODIGO UNSPSC",
        "DESCRIPCION UNSPSC",
        "FECHA DE INICIO DEL CONTRATO",
        "FECHA TERMINACION DEL CONTRATO",
        "ESTADO",
        "LUGAR DE EXPEDICIÓN",
        "CORREO",
        "TIPO DE INFORME",
        "PERIODO INFORME DESDE",
        "PERIODO INFORME HASTA",
        "PAGO No",
        "VALOR A PAGAR",
        "FECHA FIRMA",
        "OTRO SI",
        "VALOR PAGADO HISTORICO",
        "N° FOLIOS",
        "ANEXA CERTIFICACION",
        "OBSERVACIONES",
        "ACTIVIDADES",
        "PLANILLA No",
        "PERIODO COTIZADO",
        "IBC",
        "EPS NOMBRE",
        "EPS VALOR",
        "ARL NOMBRE",
        "ARL VALOR",
        "AFP NOMBRE",
        "AFP VALOR",
        "CCF NOMBRE",
        "CCF VALOR",
        "SENA VALOR",
        "ICBF VALOR",
    ]

    VERDE = "1A7A4A"
    header_fill = PatternFill("solid", fgColor=VERDE)
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=9)
    data_align = Alignment(vertical="center", wrap_text=True)

    # Escribir encabezados en fila 1
    for col, header in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 30

    # Fila de ejemplo (fila 2)
    ejemplo = [
        2025001,                          # N° DE CONTRATO
        "CARLOS ANDRES MARTINEZ ROJAS",   # NOMBRE CONTRATISTA
        "1023456789",                     # No. DE IDENTIFICACIÓN
        "BOGOTÁ",                         # EXPEDIDA EN
        "3001234567",                     # No. TELÉFONO y/o CELULAR
        "CARRERA 10 # 15-30",            # DIRECCION
        "NATURAL",                        # TIPO DE PERSONA
        "MEDICO",                         # PERFIL
        50000000,                         # VALOR TOTAL DEL CONTRATO
        48000000,                         # VALOR FINAL
        "PRESTACIÓN DE SERVICIOS PROFESIONALES "
        "EN MEDICINA GENERAL EN EL CENTRO DE SALUD "
        "DEL MUNICIPIO DURANTE LA VIGENCIA 2025",
        "MARIA FERNANDA GARCIA",          # SUPERVISOR
        "523456789",                      # CEDULA SUPERVISOR
        "PROFESIONAL ESPECIALIZADO",      # NIVEL SUPERVISOR
        "",                               # INTERVENTOR
        "",                               # NIVEL INTERVENTOR
        "CDP-2025-00123",                 # CDP No.
        "CRP-2025-00456",                 # CRP No.
        "C-1-3-02-01-001",                # IMPUTACIÓN PRESUPUESTAL
        "CENTRO DE SALUD MUNICIPAL",      # UNIDAD DE ATENCION
        "8621",                            # CODIGO CIIU
        "",                               # TIEMPO ADICION
        "MENSUAL",                        # FORMA PAGO
        "85121800",                       # CODIGO UNSPSC
        "SERVICIOS DE SALUD",             # DESCRIPCION UNSPSC
        "2025-01-15",                     # FECHA DE INICIO DEL CONTRATO
        "2025-12-31",                     # FECHA TERMINACION DEL CONTRATO
        "ACTIVO",                         # ESTADO
        "BOGOTÁ",                         # LUGAR DE EXPEDICIÓN
        "carlos.martinez@email.com",      # CORREO
        "SUPERVISION",                    # TIPO DE INFORME
        "2025-01-15",                     # PERIODO INFORME DESDE
        "2025-01-31",                     # PERIODO INFORME HASTA
        1,                                 # PAGO No
        4000000,                          # VALOR A PAGAR
        "2025-01-15",                     # FECHA FIRMA
        0,                                 # OTRO SI
        0,                                 # VALOR PAGADO HISTORICO
        "5",                              # N° FOLIOS
        "SI",                              # ANEXA CERTIFICACION
        "PAGO CORRESPONDIENTE AL MES DE ENERO",  # OBSERVACIONES
        "ACTIVIDADES DE SUPERVISION MEDICA",     # ACTIVIDADES
        "PL-2025-001",                    # PLANILLA No
        "ENERO 2025",                     # PERIODO COTIZADO
        1400000,                          # IBC
        "NUEVA EPS",                      # EPS NOMBRE
        133000,                           # EPS VALOR
        "ARL SURA",                       # ARL NOMBRE
        9660,                             # ARL VALOR
        "PORVENIR",                       # AFP NOMBRE
        224000,                           # AFP VALOR
        "COMPENSAR",                      # CCF NOMBRE
        56000,                            # CCF VALOR
        28000,                            # SENA VALOR
        14000,                            # ICBF VALOR
    ]

    # Columnas que deben tener formato numérico
    columnas_numericas = {9, 10, 35, 38, 45, 47, 49, 51, 53, 54, 55}

    for col, value in enumerate(ejemplo, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = data_font
        cell.alignment = data_align
        if col in columnas_numericas and isinstance(value, (int, float)):
            cell.number_format = '#,##0'
    ws.row_dimensions[2].height = 40

    # Ancho de columnas
    anchos = {
        1: 16, 2: 32, 3: 18, 4: 14, 5: 22, 6: 28, 7: 16, 8: 20,
        9: 20, 10: 18, 11: 50, 12: 28, 13: 18, 14: 26, 15: 28, 16: 24,
        17: 14, 18: 14, 19: 24, 20: 28, 21: 14, 22: 16, 23: 18,
        24: 16, 25: 26, 26: 24, 27: 24, 28: 12, 29: 16, 30: 30,
        31: 18, 32: 20, 33: 20, 34: 12, 35: 16, 36: 16, 37: 12,
        38: 20, 39: 12, 40: 18, 41: 40, 42: 40, 43: 16, 44: 18,
        45: 14, 46: 22, 47: 14, 48: 20, 49: 14, 50: 20, 51: 14,
        52: 20, 53: 14, 54: 14, 55: 14,
    }
    for col, width in anchos.items():
        col_letter = chr(64 + col) if col <= 26 else "A"  # fallback, no esperamos > 55
        # Usar número de columna para openpyxl
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_importacion_contratos.xlsx"},
    )
