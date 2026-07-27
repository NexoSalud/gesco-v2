"""Router para Apoyo Administrativo — CRUD, actividades y evaluación."""

import io
import logging
import os
import uuid
import base64
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy import select, func, case as sql_case
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.apoyo_administrativo import ApoyoAdministrativo
from app.models.actividad_apoyo import ActividadApoyo
from app.models.evidencia_apoyo import EvidenciaApoyo
from app.schemas.apoyo_administrativo import (
    ApoyoOut, ApoyoCreate, ApoyoUpdate,
    ActividadApoyoOut, ActividadApoyoCreate,
    EvidenciaApoyoOut, EvidenciaApoyoEvaluar,
    DashboardApoyo, ResumenApoyo,
)
from app.routers.auth import get_current_user
from app.models.auth import Usuario

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/apoyo", tags=["Apoyo Administrativo"])

EVIDENCIAS_DIR = "/app/uploads/evidencias_apoyo"
os.makedirs(EVIDENCIAS_DIR, exist_ok=True)


# ─── CRUD Apoyo Administrativo ─────────────────────────────────────────────

@router.get("", response_model=list[ApoyoOut])
@router.get("/", response_model=list[ApoyoOut])
async def listar_apoyos(
    buscar: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    stmt = select(ApoyoAdministrativo).order_by(ApoyoAdministrativo.nombre)
    if buscar:
        stmt = stmt.where(
            ApoyoAdministrativo.nombre.ilike(f"%{buscar}%") |
            ApoyoAdministrativo.identificacion.ilike(f"%{buscar}%")
        )
    result = await db.execute(stmt)
    return result.scalars().all()


@router.get("/informe-mensual")
async def descargar_informe_mensual(
    mes: int = Query(..., ge=1, le=12, description="Número del mes (1-12)"),
    anio: int = Query(..., ge=2020, description="Año"),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Descarga el informe mensual de actividades de Apoyo Administrativo en DOCX."""
    from app.services.informe_apoyo_docx import generar_informe_apoyo
    from fastapi.responses import StreamingResponse

    buf = await generar_informe_apoyo(db, mes, anio)
    mes_nombre = [
        "", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
        "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
    ][mes]
    filename = f"INFORME_ACTIVIDADES_APOYO_ADVO_EBS_{mes_nombre}_{anio}.docx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get("/{apoyo_id}", response_model=ApoyoOut)
async def obtener_apoyo(
    apoyo_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    result = await db.execute(select(ApoyoAdministrativo).where(ApoyoAdministrativo.id == apoyo_id))
    apoyo = result.scalar_one_or_none()
    if not apoyo:
        raise HTTPException(404, "Apoyo administrativo no encontrado")
    return apoyo


@router.post("", response_model=ApoyoOut, status_code=201)
@router.post("/", response_model=ApoyoOut, status_code=201)
async def crear_apoyo(
    data: ApoyoCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    # Verificar identificación única
    result = await db.execute(
        select(ApoyoAdministrativo).where(ApoyoAdministrativo.identificacion == data.identificacion)
    )
    if result.scalar_one_or_none():
        raise HTTPException(400, "Ya existe un apoyo administrativo con esa identificación")

    apoyo = ApoyoAdministrativo(**data.model_dump())
    db.add(apoyo)
    await db.commit()
    await db.refresh(apoyo)
    return apoyo


@router.put("/{apoyo_id}", response_model=ApoyoOut)
async def actualizar_apoyo(
    apoyo_id: int,
    data: ApoyoUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    result = await db.execute(select(ApoyoAdministrativo).where(ApoyoAdministrativo.id == apoyo_id))
    apoyo = result.scalar_one_or_none()
    if not apoyo:
        raise HTTPException(404, "Apoyo administrativo no encontrado")

    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(apoyo, key, val)

    await db.commit()
    await db.refresh(apoyo)
    return apoyo


# ─── Actividades ───────────────────────────────────────────────────────────

@router.get("/{apoyo_id}/actividades", response_model=list[ActividadApoyoOut])
async def listar_actividades(
    apoyo_id: int,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ActividadApoyo)
        .where(ActividadApoyo.apoyo_id == apoyo_id)
        .order_by(ActividadApoyo.orden)
    )
    return result.scalars().all()


@router.post("/{apoyo_id}/actividades", response_model=ActividadApoyoOut, status_code=201)
async def crear_actividad(
    apoyo_id: int,
    data: ActividadApoyoCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    result = await db.execute(select(ApoyoAdministrativo).where(ApoyoAdministrativo.id == apoyo_id))
    if not result.scalar_one_or_none():
        raise HTTPException(404, "Apoyo administrativo no encontrado")

    act = ActividadApoyo(apoyo_id=apoyo_id, **data.model_dump())
    db.add(act)
    await db.commit()
    await db.refresh(act)
    return act


@router.delete("/actividades/{actividad_id}", status_code=204)
async def eliminar_actividad(
    actividad_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    result = await db.execute(select(ActividadApoyo).where(ActividadApoyo.id == actividad_id))
    act = result.scalar_one_or_none()
    if not act:
        raise HTTPException(404, "Actividad no encontrada")
    await db.delete(act)
    await db.commit()
    return None


# ─── Evaluación (Dashboard + evidencias) ───────────────────────────────────

@router.get("/evaluacion/buscar", response_model=DashboardApoyo)
async def buscar_apoyo(
    cedula: str = Query(..., min_length=1),
    db: AsyncSession = Depends(get_db),
):
    """Busca un apoyo administrativo por cédula y devuelve sus actividades con evidencias."""
    result = await db.execute(
        select(ApoyoAdministrativo).where(ApoyoAdministrativo.identificacion == cedula)
    )
    apoyo = result.scalar_one_or_none()
    if not apoyo:
        raise HTTPException(404, "Apoyo administrativo no encontrado con esa cédula")

    acts_result = await db.execute(
        select(ActividadApoyo)
        .options(selectinload(ActividadApoyo.evidencias_apoyo))
        .where(ActividadApoyo.apoyo_id == apoyo.id)
        .order_by(ActividadApoyo.orden)
    )
    actividades = acts_result.scalars().all()

    actividades_data = []
    for act in actividades:
        evidencias_out = []
        for ev in act.evidencias_apoyo:
            evidencias_out.append({
                "id": ev.id,
                "actividad_apoyo_id": ev.actividad_apoyo_id,
                "apoyo_id": ev.apoyo_id,
                "tipo": ev.tipo,
                "contenido_texto": ev.contenido_texto,
                "archivo_ruta": ev.archivo_ruta,
                "archivo_nombre": ev.archivo_nombre,
                "archivo_tipo": ev.archivo_tipo,
                "estado": ev.estado,
                "observacion_coordinadora": ev.observacion_coordinadora,
                "created_at": str(ev.created_at) if ev.created_at else None,
                "evaluated_at": str(ev.evaluated_at) if ev.evaluated_at else None,
                "actividad_descripcion": act.descripcion,
            })

        actividades_data.append({
            "id": act.id,
            "descripcion": act.descripcion,
            "tipo": act.tipo,
            "orden": act.orden,
            "evidencias": evidencias_out,
        })

    return DashboardApoyo(
        apoyo_id=apoyo.id,
        identificacion=apoyo.identificacion,
        nombre=apoyo.nombre,
        telefono=apoyo.telefono,
        correo=apoyo.correo,
        perfil=apoyo.perfil,
        actividades=actividades_data,
    )


@router.get("/evaluacion/listar", response_model=list[dict])
async def listar_apoyos_evaluacion(
    buscar: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Lista apoyos administrativos con información de evidencias."""
    stmt = (
        select(
            ApoyoAdministrativo.id,
            ApoyoAdministrativo.identificacion,
            ApoyoAdministrativo.nombre,
            ApoyoAdministrativo.telefono,
            ApoyoAdministrativo.correo,
            ApoyoAdministrativo.perfil,
            func.count(EvidenciaApoyo.id).label("total_evidencias"),
            func.sum(
                sql_case((EvidenciaApoyo.estado == "PENDIENTE", 1), else_=0),
            ).label("pendientes"),
        )
        .outerjoin(ActividadApoyo, ActividadApoyo.apoyo_id == ApoyoAdministrativo.id)
        .outerjoin(EvidenciaApoyo, EvidenciaApoyo.apoyo_id == ApoyoAdministrativo.id)
        .group_by(ApoyoAdministrativo.id)
        .order_by(ApoyoAdministrativo.nombre)
    )
    if buscar:
        stmt = stmt.where(
            ApoyoAdministrativo.nombre.ilike(f"%{buscar}%") |
            ApoyoAdministrativo.identificacion.ilike(f"%{buscar}%")
        )

    result = await db.execute(stmt)
    rows = result.all()

    return [
        {
            "id": row.id,
            "identificacion": row.identificacion,
            "nombre": row.nombre,
            "telefono": row.telefono,
            "correo": row.correo,
            "perfil": row.perfil,
            "total_evidencias": row.total_evidencias or 0,
            "pendientes": row.pendientes or 0,
            "tipo": "APOYO",
        }
        for row in rows
    ]


@router.post("/evidencias", response_model=EvidenciaApoyoOut, status_code=201)
async def subir_evidencia(
    actividad_apoyo_id: int = Form(...),
    apoyo_id: int = Form(...),
    tipo: str = Form(...),
    contenido_texto: str | None = Form(None),
    archivo: UploadFile | None = File(None),
    db: AsyncSession = Depends(get_db),
):
    """Sube una evidencia (archivo, imagen o texto) para una actividad de apoyo."""
    if tipo not in ("ARCHIVO", "TEXTO", "IMAGEN"):
        raise HTTPException(400, "Tipo debe ser ARCHIVO, TEXTO o IMAGEN")

    # Validar actividad
    result = await db.execute(
        select(ActividadApoyo).where(
            ActividadApoyo.id == actividad_apoyo_id,
            ActividadApoyo.apoyo_id == apoyo_id,
        )
    )
    actividad = result.scalar_one_or_none()
    if not actividad:
        raise HTTPException(404, "Actividad no encontrada para ese apoyo")

    # Validar apoyo
    result = await db.execute(select(ApoyoAdministrativo).where(ApoyoAdministrativo.id == apoyo_id))
    if not result.scalar_one_or_none():
        raise HTTPException(404, "Apoyo administrativo no encontrado")

    ev_data = {
        "actividad_apoyo_id": actividad_apoyo_id,
        "apoyo_id": apoyo_id,
        "tipo": tipo,
        "contenido_texto": contenido_texto,
    }

    if tipo in ("ARCHIVO", "IMAGEN") and archivo:
        content = await archivo.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(400, "El archivo excede el tamaño máximo de 10MB")

        ext = ""
        if archivo.filename and "." in archivo.filename:
            ext = archivo.filename.rsplit(".", 1)[-1]
        safe_name = f"{uuid.uuid4()}.{ext}"
        file_path = os.path.join(EVIDENCIAS_DIR, safe_name)
        with open(file_path, "wb") as f:
            f.write(content)

        ev_data["archivo_ruta"] = f"/uploads/evidencias_apoyo/{safe_name}"
        ev_data["archivo_nombre"] = archivo.filename
        ev_data["archivo_tipo"] = archivo.content_type

    elif tipo == "TEXTO" and not contenido_texto:
        raise HTTPException(400, "Para tipo TEXTO debe proporcionar contenido_texto")

    ev = EvidenciaApoyo(**ev_data)
    db.add(ev)
    await db.commit()
    await db.refresh(ev)

    return EvidenciaApoyoOut(
        id=ev.id,
        actividad_apoyo_id=ev.actividad_apoyo_id,
        apoyo_id=ev.apoyo_id,
        tipo=ev.tipo,
        contenido_texto=ev.contenido_texto,
        archivo_ruta=ev.archivo_ruta,
        archivo_nombre=ev.archivo_nombre,
        archivo_tipo=ev.archivo_tipo,
        estado=ev.estado,
        observacion_coordinadora=ev.observacion_coordinadora,
        created_at=ev.created_at,
        evaluated_at=ev.evaluated_at,
        evaluated_by=ev.evaluated_by,
        actividad_descripcion=actividad.descripcion,
    )


@router.put("/evidencias/{evidencia_id}", response_model=EvidenciaApoyoOut)
async def evaluar_evidencia(
    evidencia_id: int,
    data: EvidenciaApoyoEvaluar,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Aprobar o rechazar una evidencia de apoyo."""
    result = await db.execute(select(EvidenciaApoyo).where(EvidenciaApoyo.id == evidencia_id))
    ev = result.scalar_one_or_none()
    if not ev:
        raise HTTPException(404, "Evidencia no encontrada")

    ev.estado = data.estado
    ev.observacion_coordinadora = data.observacion
    ev.evaluated_by = current_user.id
    ev.evaluated_at = datetime.utcnow()

    await db.commit()
    await db.refresh(ev)

    act_result = await db.execute(
        select(ActividadApoyo.descripcion).where(ActividadApoyo.id == ev.actividad_apoyo_id)
    )
    act_desc = act_result.scalar_one_or_none()

    return EvidenciaApoyoOut(
        id=ev.id,
        actividad_apoyo_id=ev.actividad_apoyo_id,
        apoyo_id=ev.apoyo_id,
        tipo=ev.tipo,
        contenido_texto=ev.contenido_texto,
        archivo_ruta=ev.archivo_ruta,
        archivo_nombre=ev.archivo_nombre,
        archivo_tipo=ev.archivo_tipo,
        estado=ev.estado,
        observacion_coordinadora=ev.observacion_coordinadora,
        created_at=ev.created_at,
        evaluated_at=ev.evaluated_at,
        evaluated_by=ev.evaluated_by,
        actividad_descripcion=act_desc,
    )


@router.get("/{apoyo_id}/resumen", response_model=ResumenApoyo)
async def resumen_apoyo(
    apoyo_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    result = await db.execute(select(ApoyoAdministrativo).where(ApoyoAdministrativo.id == apoyo_id))
    apoyo = result.scalar_one_or_none()
    if not apoyo:
        raise HTTPException(404, "Apoyo no encontrado")

    # Total de actividades
    act_result = await db.execute(
        select(func.count(ActividadApoyo.id)).where(ActividadApoyo.apoyo_id == apoyo_id)
    )
    total_act = act_result.scalar() or 0

    # Evidencias por estado
    ev_result = await db.execute(
        select(EvidenciaApoyo.estado, func.count(EvidenciaApoyo.id))
        .where(EvidenciaApoyo.apoyo_id == apoyo_id)
        .group_by(EvidenciaApoyo.estado)
    )
    counts = {row[0]: row[1] for row in ev_result.all()}

    # Actividades con al menos una evidencia
    act_con_ev = await db.execute(
        select(func.count(func.distinct(EvidenciaApoyo.actividad_apoyo_id)))
        .where(EvidenciaApoyo.apoyo_id == apoyo_id)
    )
    con_ev = act_con_ev.scalar() or 0

    aprobadas = counts.get("APROBADO", 0)
    rechazadas = counts.get("RECHAZADO", 0)
    pendientes = counts.get("PENDIENTE", 0)
    total_ev = aprobadas + rechazadas + pendientes
    porcentaje = round(aprobadas / total_ev * 100, 1) if total_ev > 0 else 0

    return ResumenApoyo(
        apoyo_id=apoyo_id,
        apoyo_nombre=apoyo.nombre,
        total_actividades=total_act,
        con_evidencia=con_ev,
        sin_evidencia=total_act - con_ev,
        aprobadas=aprobadas,
        rechazadas=rechazadas,
        pendientes=pendientes,
        porcentaje_cumplimiento=porcentaje,
    )

# ─── Endpoint de carga masiva desde el seed_data (una sola vez) ────────────




@router.post("/seed-actividades", status_code=200)
async def seed_actividades(
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Crea actividades desde el DOCX para apoyos que no tengan ninguna."""
    from app.seed_apoyo import SECTIONS

    creadas = 0
    omitidas = 0

    for section in SECTIONS:
        for name in section["names"]:
            result = await db.execute(
                select(ApoyoAdministrativo).where(ApoyoAdministrativo.nombre == name)
            )
            apoyo = result.scalar_one_or_none()
            if not apoyo:
                logger.warning(f"Apoyo no encontrado: {name}")
                continue

            act_result = await db.execute(
                select(ActividadApoyo).where(ActividadApoyo.apoyo_id == apoyo.id).limit(1)
            )
            if act_result.scalar_one_or_none():
                omitidas += len(section["actividades"])
                continue

            for i, at in enumerate(section["actividades"]):
                db.add(ActividadApoyo(apoyo_id=apoyo.id, descripcion=at, tipo="GENERAL", orden=i + 1))
            creadas += len(section["actividades"])

    await db.commit()
    return {"actividades_creadas": creadas, "actividades_omitidas": omitidas}


@router.post("/importar-excel", status_code=200)
async def importar_apoyos_excel(
    archivo: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Importa apoyos administrativos y sus actividades desde un archivo Excel.

    Formato esperado (columnas):
        PERFIL | NOMBRE_COMPLETO | ORDEN | ACTIVIDAD

    - Si el apoyo ya existe (por nombre exacto), se reemplazan todas sus actividades.
    - Si no existe, se crea con identificación autogenerada.
    - Las actividades se asignan en el orden indicado.
    """
    if not archivo.filename or not archivo.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Debe subir un archivo Excel (.xlsx o .xls)")

    import openpyxl

    content = await archivo.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    if ws is None:
        raise HTTPException(400, "El archivo Excel está vacío")

    # Validate headers
    headers = [str(c.value).strip().upper() if c.value else "" for c in ws[1]]
    required = {"PERFIL", "NOMBRE_COMPLETO", "ORDEN", "ACTIVIDAD"}
    if not required.issubset(set(headers)):
        raise HTTPException(400, f"Columnas requeridas: PERFIL, NOMBRE_COMPLETO, ORDEN, ACTIVIDAD. Encontradas: {headers}")

    # Map column indices
    col_map = {}
    for i, h in enumerate(headers):
        col_map[h] = i

    # Read all rows (skip header)
    groups = {}  # key: (perfil_lower, nombre_lower) -> {perfil, nombre, actividades: [(orden, texto)]}
    for row in ws.iter_rows(min_row=2, values_only=False):
        perfil = str(row[col_map["PERFIL"]].value or "").strip()
        nombre = str(row[col_map["NOMBRE_COMPLETO"]].value or "").strip()
        orden_val = row[col_map["ORDEN"]].value
        actividad = str(row[col_map["ACTIVIDAD"]].value or "").strip()

        if not perfil or not nombre or not actividad:
            continue

        try:
            orden = int(orden_val) if orden_val else 0
        except (ValueError, TypeError):
            orden = 0

        key = (perfil.upper(), nombre.upper())
        if key not in groups:
            groups[key] = {
                "perfil": perfil,
                "nombre": nombre,
                "actividades": [],
            }
        groups[key]["actividades"].append((orden, actividad))

    if not groups:
        raise HTTPException(400, "No se encontraron datos válidos en el archivo")

    creados = 0
    actualizados = 0
    total_actividades = 0

    for key, group in groups.items():
        perfil = group["perfil"]
        nombre = group["nombre"]
        actividades = sorted(group["actividades"], key=lambda x: x[0])

        # Buscar apoyo existente por nombre exacto (case-insensitive)
        result = await db.execute(
            select(ApoyoAdministrativo).where(
                func.upper(ApoyoAdministrativo.nombre) == nombre.upper()
            )
        )
        apoyo = result.scalar_one_or_none()

        if apoyo:
            # Actualizar perfil si cambió
            if perfil and apoyo.perfil != perfil:
                apoyo.perfil = perfil
            actualizados += 1
        else:
            # Crear nuevo apoyo con identificación autogenerada
            apoyo = ApoyoAdministrativo(
                nombre=nombre,
                identificacion=f"APOYO-{nombre.split()[0][:10].upper()}-{len(nombre):03d}",
                telefono=None,
                correo=None,
                perfil=perfil,
                activo=True,
            )
            db.add(apoyo)
            await db.flush()  # Obtener ID
            creados += 1

        # Eliminar actividades existentes
        existing = await db.execute(
            select(ActividadApoyo).where(ActividadApoyo.apoyo_id == apoyo.id)
        )
        for act in existing.scalars().all():
            await db.delete(act)

        # Crear nuevas actividades
        for orden, descripcion in actividades:
            db.add(ActividadApoyo(
                apoyo_id=apoyo.id,
                descripcion=descripcion,
                tipo="GENERAL",
                orden=orden,
            ))
            total_actividades += 1

    await db.commit()

    return {
        "mensaje": "Importación completada",
        "creados": creados,
        "actualizados": actualizados,
        "total_actividades": total_actividades,
        "perfiles_cargados": len(groups),
    }
