"""Router para la gestión de inventario, almacenes, movimientos y actas."""

import io
import os
import logging
from datetime import datetime, date
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, Response
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy import select, func, delete
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Optional

from app.database import get_db
from app.config import settings
from app.models.resolucion import Resolucion
from app.models.contrato import Contrato
from app.models.almacen import Almacen
from app.models.inventario import Articulo, UnidadInventario, MovimientoInventario, Acta
from app.models.auth import Usuario
from app.routers.auth import get_current_user
from app.schemas.inventario import (
    AlmacenCreate, AlmacenUpdate, AlmacenResponse,
    ArticuloCreate, ArticuloResponse, ArticuloUpdate,
    UnidadInventarioCreate, UnidadInventarioResponse, UnidadInventarioUpdate,
    EntregaRequest, DevolucionRequest,
    ActaResponse, ActaListResponse, MovimientoResponse,
    ImportInventarioResult, DashboardResponse,
    CategoriaResumen, ArticuloDashboardInfo,
    BulkDeleteRequest
)
from app.services.inventory_docx_generator import generar_acta_inventario_docx

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/inventario", tags=["Inventario"])

ESTADOS_DISPONIBLES = {"DISPONIBLE", "BUEN_ESTADO", "REGULAR"}


# ─── MÉTODOS AUXILIARES PARA EXCEL ───
def _clean_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _parse_int(val) -> int:
    if val is None:
        return 0
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return 0


def _normalize_header(title: str) -> str:
    # Quitar tildes, espacios múltiples, caracteres raros
    t = title.strip().upper()
    t = re.sub(r"[^A-Z0-9 ]", "", t)
    return re.sub(r"\s+", " ", t)


def _find_col_idx(headers: list, target: str) -> Optional[int]:
    target_norm = _normalize_header(target)
    for i, h in enumerate(headers):
        if h and _normalize_header(str(h)) == target_norm:
            return i
    return None


import re


# ─── ENDPOINTS ALMACENES ───
@router.get("/almacenes", response_model=List[AlmacenResponse])
async def listar_almacenes(db: AsyncSession = Depends(get_db)):
    """Lista todos los almacenes registrados."""
    # Pre-seed de almacén por defecto si no hay ninguno
    res = await db.execute(select(Almacen))
    almacenes = res.scalars().all()
    if not almacenes:
        default_almacen = Almacen(nombre="Almacén Principal", ubicacion="Sede Principal", responsable="Coordinador de Apoyo")
        db.add(default_almacen)
        await db.commit()
        res = await db.execute(select(Almacen))
        almacenes = res.scalars().all()
    return almacenes


@router.post("/almacenes", response_model=AlmacenResponse)
async def crear_almacen(almacen: AlmacenCreate, db: AsyncSession = Depends(get_db)):
    """Crea un nuevo almacén."""
    existing = await db.execute(select(Almacen).where(Almacen.nombre == almacen.nombre))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Ya existe un almacén con este nombre")
    
    new_almacen = Almacen(**almacen.model_dump())
    db.add(new_almacen)
    await db.commit()
    await db.refresh(new_almacen)
    return new_almacen


@router.put("/almacenes/{almacen_id}", response_model=AlmacenResponse)
async def actualizar_almacen(
    almacen_id: int, 
    almacen: AlmacenUpdate, 
    db: AsyncSession = Depends(get_db)
):
    """Actualiza los datos de un almacén existente."""
    res = await db.execute(select(Almacen).where(Almacen.id == almacen_id))
    db_almacen = res.scalar_one_or_none()
    if not db_almacen:
        raise HTTPException(status_code=404, detail="Almacén no encontrado")
        
    update_data = almacen.model_dump(exclude_unset=True)
    if "nombre" in update_data and update_data["nombre"] != db_almacen.nombre:
        existing = await db.execute(
            select(Almacen).where(Almacen.nombre == update_data["nombre"])
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Ya existe otro almacén con este nombre")

    for key, value in update_data.items():
        setattr(db_almacen, key, value)
        
    await db.commit()
    await db.refresh(db_almacen)
    return db_almacen


@router.delete("/almacenes/{almacen_id}")
async def eliminar_almacen(almacen_id: int, db: AsyncSession = Depends(get_db)):
    """Elimina un almacén si no está referenciado por artículos ni unidades físicas."""
    res = await db.execute(select(Almacen).where(Almacen.id == almacen_id))
    db_almacen = res.scalar_one_or_none()
    if not db_almacen:
        raise HTTPException(status_code=404, detail="Almacén no encontrado")
        
    # Verificar referencias en artículos
    res_art = await db.execute(select(Articulo).where(Articulo.almacen_id == almacen_id))
    if res_art.scalars().first():
        raise HTTPException(
            status_code=400, 
            detail="No se puede eliminar el almacén porque contiene artículos en el catálogo"
        )
        
    # Verificar referencias en unidades de inventario
    res_unidades = await db.execute(
        select(UnidadInventario).where(UnidadInventario.almacen_id == almacen_id)
    )
    if res_unidades.scalars().first():
        raise HTTPException(
            status_code=400, 
            detail="No se puede eliminar el almacén porque tiene unidades físicas de inventario asignadas"
        )
        
    await db.delete(db_almacen)
    await db.commit()
    return {"message": "Almacén eliminado correctamente"}


# ─── ENDPOINTS ARTÍCULOS Y UNIDADES ───
@router.get("/articulos", response_model=List[ArticuloResponse])
async def listar_articulos(categoria: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """Lista el catálogo de artículos."""
    q = select(Articulo)
    if categoria:
        q = q.where(Articulo.categoria == categoria)
    res = await db.execute(q.order_by(Articulo.elemento))
    return res.scalars().all()


@router.post("/articulo", response_model=ArticuloResponse)
async def crear_articulo(articulo: ArticuloCreate, db: AsyncSession = Depends(get_db)):
    """Crea manualmente un artículo en el catálogo maestro."""
    # Buscar si ya existe
    q = select(Articulo).where(
        Articulo.categoria == articulo.categoria,
        Articulo.elemento == articulo.elemento
    )
    if articulo.marca:
        q = q.where(Articulo.marca == articulo.marca)
    if articulo.modelo:
        q = q.where(Articulo.modelo == articulo.modelo)
        
    existing = await db.execute(q)
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Ya existe este artículo en el catálogo")

    # Si es dotación/insumo no requiere serial
    req_serial = articulo.requiere_serial
    if articulo.categoria == "DOTACION_INSUMO":
        req_serial = False

    new_articulo = Articulo(
        categoria=articulo.categoria,
        tipo_elemento=articulo.tipo_elemento.upper(),
        elemento=articulo.elemento,
        marca=articulo.marca,
        modelo=articulo.modelo,
        requiere_serial=req_serial,
        stock_total=articulo.stock_total or 0,
        stock_disponible=articulo.stock_total or 0,
        almacen_id=articulo.almacen_id,
        resolucion_id=articulo.resolucion_id
    )
    db.add(new_articulo)
    await db.commit()
    await db.refresh(new_articulo)
    return new_articulo


@router.get("/unidades", response_model=List[UnidadInventarioResponse])
async def listar_unidades(
    articulo_id: Optional[int] = None,
    estado: Optional[str] = None,
    categoria: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """Lista las unidades físicas (tecnológicas y biomédicas)."""
    q = select(UnidadInventario).options(
        selectinload(UnidadInventario.articulo),
        selectinload(UnidadInventario.contrato_actual).selectinload(Contrato.contratista_rel)
    ).join(Articulo)
    if articulo_id:
        q = q.where(UnidadInventario.articulo_id == articulo_id)
    if estado:
        q = q.where(UnidadInventario.estado == estado)
    if categoria:
        q = q.where(Articulo.categoria == categoria)
        
    res = await db.execute(q.order_by(UnidadInventario.serial))
    return res.scalars().all()


@router.post("/unidad", response_model=UnidadInventarioResponse)
async def crear_unidad(unidad: UnidadInventarioCreate, db: AsyncSession = Depends(get_db)):
    """Registra una unidad individual en inventario."""
    # Verificar artículo
    res_art = await db.execute(select(Articulo).where(Articulo.id == unidad.articulo_id))
    articulo = res_art.scalar_one_or_none()
    if not articulo:
        raise HTTPException(status_code=404, detail="Artículo no encontrado en catálogo")
    if not articulo.requiere_serial:
        raise HTTPException(status_code=400, detail="Este artículo se maneja por cantidad, no requiere unidades serializadas")

    # Verificar serial único
    if unidad.serial:
        res_un = await db.execute(select(UnidadInventario).where(UnidadInventario.serial == unidad.serial))
        if res_un.scalar_one_or_none():
            raise HTTPException(status_code=400, detail=f"Ya existe una unidad registrada con el serial {unidad.serial}")

    # Si no se define almacén, tomar el primero
    almacen_id = unidad.almacen_id
    if not almacen_id:
        res_alm = await db.execute(select(Almacen.id))
        almacen_id = res_alm.scalars().first()

    new_unidad = UnidadInventario(
        articulo_id=unidad.articulo_id,
        almacen_id=almacen_id,
        serial=unidad.serial,
        imei2=unidad.imei2,
        estado=unidad.estado or "DISPONIBLE",
        observaciones=unidad.observaciones,
        resolucion_id=unidad.resolucion_id
    )
    db.add(new_unidad)
    
    # Incrementar stock del artículo maestro
    articulo.stock_total += 1
    if new_unidad.estado in ESTADOS_DISPONIBLES:
        articulo.stock_disponible += 1
        
    await db.commit()
    res_unit = await db.execute(
        select(UnidadInventario)
        .options(
            selectinload(UnidadInventario.articulo),
            selectinload(UnidadInventario.contrato_actual).selectinload(Contrato.contratista_rel)
        )
        .where(UnidadInventario.id == new_unidad.id)
    )
    return res_unit.scalar_one()


@router.put("/articulo/{articulo_id}", response_model=ArticuloResponse)
async def actualizar_articulo(articulo_id: int, articulo_data: ArticuloUpdate, db: AsyncSession = Depends(get_db)):
    """Actualiza la información de un artículo."""
    res = await db.execute(select(Articulo).where(Articulo.id == articulo_id))
    db_articulo = res.scalar_one_or_none()
    if not db_articulo:
        raise HTTPException(status_code=404, detail="Artículo no encontrado")
        
    for key, value in articulo_data.model_dump(exclude_unset=True).items():
        if key == "stock_total":
            if not db_articulo.requiere_serial:
                diff = value - db_articulo.stock_total
                db_articulo.stock_total = value
                db_articulo.stock_disponible += diff
        else:
            setattr(db_articulo, key, value)
            
    await db.commit()
    await db.refresh(db_articulo)
    return db_articulo


@router.delete("/articulo/{articulo_id}")
async def eliminar_articulo(articulo_id: int, db: AsyncSession = Depends(get_db)):
    """Elimina un artículo y sus unidades asociadas."""
    res = await db.execute(select(Articulo).where(Articulo.id == articulo_id))
    db_articulo = res.scalar_one_or_none()
    if not db_articulo:
        raise HTTPException(status_code=404, detail="Artículo no encontrado")
        
    # Eliminar unidades de inventario asociadas
    await db.execute(
        delete(UnidadInventario).where(UnidadInventario.articulo_id == articulo_id)
    )
    # Eliminar movimientos asociados
    await db.execute(
        delete(MovimientoInventario).where(MovimientoInventario.articulo_id == articulo_id)
    )
    # Eliminar el artículo
    await db.delete(db_articulo)
    await db.commit()
    return {"message": "Artículo y sus elementos asociados eliminados correctamente"}


@router.post("/articulos/bulk-delete")
async def eliminar_articulos_masivo(req: BulkDeleteRequest, db: AsyncSession = Depends(get_db)):
    """Elimina varios artículos de manera masiva."""
    if not req.ids:
        return {"message": "No se proporcionaron IDs"}
        
    # Eliminar unidades de inventario asociadas
    await db.execute(
        delete(UnidadInventario).where(UnidadInventario.articulo_id.in_(req.ids))
    )
    # Eliminar movimientos asociados
    await db.execute(
        delete(MovimientoInventario).where(MovimientoInventario.articulo_id.in_(req.ids))
    )
    # Eliminar artículos
    await db.execute(
        delete(Articulo).where(Articulo.id.in_(req.ids))
    )
    await db.commit()
    return {"message": "Artículos eliminados masivamente de manera correcta"}


@router.put("/unidad/{unidad_id}", response_model=UnidadInventarioResponse)
async def actualizar_unidad(unidad_id: int, unidad_data: UnidadInventarioUpdate, db: AsyncSession = Depends(get_db)):
    """Actualiza la información de una unidad física."""
    res = await db.execute(
        select(UnidadInventario)
        .options(
            selectinload(UnidadInventario.articulo),
            selectinload(UnidadInventario.contrato_actual).selectinload(Contrato.contratista_rel)
        )
        .where(UnidadInventario.id == unidad_id)
    )
    db_unidad = res.scalar_one_or_none()
    if not db_unidad:
        raise HTTPException(status_code=404, detail="Unidad física no encontrada")
        
    old_estado = db_unidad.estado
    
    for key, value in unidad_data.model_dump(exclude_unset=True).items():
        setattr(db_unidad, key, value)
        
    # Si cambió el estado, reajustar stock_disponible del artículo maestro
    if unidad_data.estado and unidad_data.estado != old_estado:
        articulo = db_unidad.articulo
        if articulo:
            was_available = old_estado in ESTADOS_DISPONIBLES
            is_available = unidad_data.estado in ESTADOS_DISPONIBLES
            if is_available and not was_available:
                articulo.stock_disponible += 1
            elif was_available and not is_available:
                articulo.stock_disponible -= 1
                
    await db.commit()
    res_reload = await db.execute(
        select(UnidadInventario)
        .options(
            selectinload(UnidadInventario.articulo),
            selectinload(UnidadInventario.contrato_actual).selectinload(Contrato.contratista_rel)
        )
        .where(UnidadInventario.id == unidad_id)
    )
    return res_reload.scalar_one()


@router.delete("/unidad/{unidad_id}")
async def eliminar_unidad(unidad_id: int, db: AsyncSession = Depends(get_db)):
    """Elimina una unidad física y actualiza el stock del artículo."""
    res = await db.execute(
        select(UnidadInventario)
        .options(selectinload(UnidadInventario.articulo))
        .where(UnidadInventario.id == unidad_id)
    )
    db_unidad = res.scalar_one_or_none()
    if not db_unidad:
        raise HTTPException(status_code=404, detail="Unidad física no encontrada")
        
    articulo = db_unidad.articulo
    if articulo:
        articulo.stock_total = max(0, articulo.stock_total - 1)
        if db_unidad.estado in ESTADOS_DISPONIBLES:
            articulo.stock_disponible = max(0, articulo.stock_disponible - 1)
            
    # Eliminar movimientos asociados a esta unidad
    await db.execute(
        delete(MovimientoInventario).where(MovimientoInventario.unidad_id == unidad_id)
    )
    
    await db.delete(db_unidad)
    await db.commit()
    return {"message": "Unidad física eliminada correctamente"}


@router.post("/unidades/bulk-delete")
async def eliminar_unidades_masivo(req: BulkDeleteRequest, db: AsyncSession = Depends(get_db)):
    """Elimina varias unidades físicas de manera masiva."""
    if not req.ids:
        return {"message": "No se proporcionaron IDs"}
        
    # Obtener las unidades para actualizar los stocks de sus artículos
    res = await db.execute(
        select(UnidadInventario)
        .options(selectinload(UnidadInventario.articulo))
        .where(UnidadInventario.id.in_(req.ids))
    )
    unidades = res.scalars().all()
    
    for u in unidades:
        articulo = u.articulo
        if articulo:
            articulo.stock_total = max(0, articulo.stock_total - 1)
            if u.estado in ESTADOS_DISPONIBLES:
                articulo.stock_disponible = max(0, articulo.stock_disponible - 1)
                
    # Eliminar movimientos asociados
    await db.execute(
        delete(MovimientoInventario).where(MovimientoInventario.unidad_id.in_(req.ids))
    )
    # Eliminar unidades
    await db.execute(
        delete(UnidadInventario).where(UnidadInventario.id.in_(req.ids))
    )
    
    await db.commit()
    return {"message": "Unidades físicas eliminadas masivamente de manera correcta"}


@router.get("/plantillas/serializado")
async def descargar_plantilla_serializado():
    """Genera y descarga la plantilla de Excel para elementos serializados."""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Serializados"
    
    # Headers
    headers = [
        "ALMACEN_ID",
        "TIPO DE ELEMENTO", 
        "ELEMENTO", 
        "MARCA", 
        "MODELO", 
        "IMEI 1 (S/N)", 
        "IMEI 2 (ID. DEL DISPOSITIVO)",
        "RESOLUCION"
    ]
    ws.append(headers)
    
    # Sample Row
    ws.append([1, "TECNOLOGICO", "Computador Portátil", "Lenovo", "Thinkpad L14", "LNV-88998877", "", "RES-2026-001"])
    ws.append([1, "BIOMEDICO", "Tensiómetro Digital", "Omron", "HEM-7120", "OMR-772211", "", "RES-2026-001"])
    
    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_inventario_serializado.xlsx"}
    )


@router.get("/plantillas/dotacion")
async def descargar_plantilla_dotacion():
    """Genera y descarga la plantilla de Excel para dotación/insumos."""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Dotacion"
    
    # Headers
    headers = ["ALMACEN_ID", "TIPO DE ELEMENTO", "ELEMENTO", "CANTIDAD", "RESOLUCION"]
    ws.append(headers)
    
    # Sample Row
    ws.append([1, "DOTACION", "Chaleco", 80, "RES-2026-001"])
    ws.append([2, "DOTACION", "Gorra", 120, "RES-2026-001"])
    
    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_inventario_dotacion.xlsx"}
    )


@router.get("/contratista/{contratista_id}")
async def detalle_inventario_contratista(contratista_id: int, db: AsyncSession = Depends(get_db)):
    """Obtiene el inventario actual (serializado y dotación) asignado a todos los contratos de un contratista."""
    from sqlalchemy.orm import selectinload
    
    # 1. Obtener contratos del contratista
    res_contratos = await db.execute(
        select(Contrato.id, Contrato.numero_contrato)
        .where(Contrato.contratista_id == contratista_id)
    )
    contratos = res_contratos.all()
    contrato_ids = [c[0] for c in contratos]
    
    if not contrato_ids:
        return {"unidades": [], "dotacion": []}

    # 2. Obtener unidades serializadas activas asignadas a cualquiera de estos contratos
    res_unidades = await db.execute(
        select(UnidadInventario)
        .options(selectinload(UnidadInventario.articulo))
        .where(UnidadInventario.contrato_actual_id.in_(contrato_ids))
    )
    unidades = res_unidades.scalars().all()

    # 3. Calcular dotación e insumos activos
    dotacion_activa = []
    insumos_activa = []
    
    for contrato_id in contrato_ids:
        res_movs = await db.execute(
            select(MovimientoInventario)
            .options(selectinload(MovimientoInventario.articulo))
            .where(MovimientoInventario.contrato_id == contrato_id)
        )
        movs = res_movs.scalars().all()
        
        balances = {}
        for m in movs:
            if m.articulo and not m.articulo.requiere_serial:
                art_id = m.articulo_id
                if art_id not in balances:
                    balances[art_id] = {
                        "articulo_id": art_id,
                        "tipo_elemento": m.articulo.tipo_elemento,
                        "elemento": m.articulo.elemento,
                        "categoria": m.articulo.categoria,
                        "entregado": 0,
                        "devuelto": 0
                    }
                if m.tipo == "ENTREGA":
                    balances[art_id]["entregado"] += m.cantidad
                elif m.tipo == "DEVOLUCION":
                    balances[art_id]["devuelto"] += m.cantidad
        
        num_contrato = next(c[1] for c in contratos if c[0] == contrato_id)
        for art_id, b in balances.items():
            saldo = b["entregado"] - b["devuelto"]
            if saldo > 0:
                item_data = {
                    "articulo_id": art_id,
                    "tipo_elemento": b["tipo_elemento"],
                    "elemento": b["elemento"],
                    "cantidad": saldo,
                    "contrato_id": contrato_id,
                    "numero_contrato": num_contrato,
                    "categoria": b["categoria"]
                }
                if b["categoria"] == "INSUMO":
                    insumos_activa.append(item_data)
                else:
                    dotacion_activa.append(item_data)

    unidades_out = []
    for u in unidades:
        num_contrato = next(c[1] for c in contratos if c[0] == u.contrato_actual_id)
        unidades_out.append({
            "id": u.id,
            "serial": u.serial,
            "imei2": u.imei2,
            "tipo_elemento": u.articulo.tipo_elemento if u.articulo else "",
            "elemento": u.articulo.elemento if u.articulo else "",
            "marca": u.articulo.marca if u.articulo else "",
            "modelo": u.articulo.modelo if u.articulo else "",
            "contrato_id": u.contrato_actual_id,
            "numero_contrato": num_contrato
        })

    return {
        "unidades": unidades_out,
        "dotacion": dotacion_activa,
        "insumos": insumos_activa
    }


# ─── ENDPOINTS IMPORTACIÓN MASIVA DESDE EXCEL ───
@router.post("/import/serializado", response_model=ImportInventarioResult)
async def importar_inventario_serializado(
    file: UploadFile = File(...),
    almacen_id: Optional[int] = Query(None),
    resolucion_id: Optional[int] = Query(None),
    dry_run: bool = Query(False),
    db: AsyncSession = Depends(get_db)
):
    """Carga masiva de elementos serializados (tecnológico/biomédico)."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "El archivo debe ser .xlsx")

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        ws = wb.active
        if ws is None:
            raise HTTPException(400, "El Excel no tiene hojas activas")
        err_col_idx = ws.max_column + 1
        ws.cell(row=1, column=err_col_idx, value="ERRORES")
    except Exception as e:
        raise HTTPException(400, f"Error al leer el Excel: {e}")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        raise HTTPException(400, "El archivo Excel está vacío")

    headers = [str(h) if h else "" for h in raw_headers]
    
    # Mapear columnas
    col_tipo = _find_col_idx(headers, "TIPO DE ELEMENTO")
    if col_tipo is None:
         col_tipo = _find_col_idx(headers, "TIPO")
    col_elemento = _find_col_idx(headers, "ELEMENTO")
    col_marca = _find_col_idx(headers, "MARCA")
    col_modelo = _find_col_idx(headers, "MODELO")
    col_serial = _find_col_idx(headers, "IMEI 1 (S/N)")
    if col_serial is None:
        col_serial = _find_col_idx(headers, "SERIAL")
    col_imei2 = _find_col_idx(headers, "IMEI 2 (ID. DEL DISPOSITIVO)")
    if col_imei2 is None:
        col_imei2 = _find_col_idx(headers, "IMEI 2")
    col_almacen = _find_col_idx(headers, "ALMACEN_ID")
    if col_almacen is None:
        col_almacen = _find_col_idx(headers, "ID ALMACEN")
    if col_almacen is None:
        col_almacen = _find_col_idx(headers, "ALMACEN")
    if col_almacen is None:
        col_almacen = _find_col_idx(headers, "ID_ALMACEN")
    col_resolucion = _find_col_idx(headers, "RESOLUCION")
    if col_resolucion is None:
        col_resolucion = _find_col_idx(headers, "CODIGO RESOLUCION")
    if col_resolucion is None:
        col_resolucion = _find_col_idx(headers, "CODIGO_RESOLUCION")
    if col_resolucion is None:
        col_resolucion = _find_col_idx(headers, "RESOLUCION_CODIGO")
    if col_resolucion is None:
        col_resolucion = _find_col_idx(headers, "RESOLUCIÓN")

    if col_tipo is None or col_elemento is None or col_serial is None or col_almacen is None:
        raise HTTPException(
            400,
            "No se encontraron las columnas requeridas en el Excel: 'TIPO DE ELEMENTO', 'ELEMENTO', 'IMEI 1 (S/N)' y 'ALMACEN_ID'."
        )

    result = ImportInventarioResult(dry_run=dry_run)
    for fila_idx, row in enumerate(rows_iter, start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        result.total += 1
        nested = await db.begin_nested()
        try:
            tipo_raw = _clean_str(row[col_tipo])
            elemento_raw = _clean_str(row[col_elemento])
            serial_raw = _clean_str(row[col_serial])
            marca_raw = _clean_str(row[col_marca]) if col_marca is not None else None
            modelo_raw = _clean_str(row[col_modelo]) if col_modelo is not None else None
            imei2_raw = _clean_str(row[col_imei2]) if col_imei2 is not None else None
            row_almacen_id = _parse_int(row[col_almacen]) if col_almacen is not None else 0
            resolucion_raw = _clean_str(row[col_resolucion]) if col_resolucion is not None else None

            if not tipo_raw or not elemento_raw or not serial_raw or not row_almacen_id:
                err_msg = "Campos obligatorios vacíos (TIPO, ELEMENTO, SERIAL o ALMACEN_ID)"
                result.errors.append({"fila": fila_idx, "error": err_msg})
                ws.cell(row=fila_idx, column=err_col_idx, value=err_msg)
                await nested.rollback()
                continue

            # Validar si el almacén existe
            res_alm = await db.execute(select(Almacen).where(Almacen.id == row_almacen_id))
            almacen_obj = res_alm.scalar_one_or_none()
            if not almacen_obj:
                err_msg = f"El almacén con ID {row_almacen_id} no existe."
                result.errors.append({"fila": fila_idx, "error": err_msg})
                ws.cell(row=fila_idx, column=err_col_idx, value=err_msg)
                await nested.rollback()
                continue

            # Buscar la resolución por código si viene en la fila
            resolucion_id_db = None
            if resolucion_raw:
                res_res = await db.execute(select(Resolucion).where(Resolucion.codigo == resolucion_raw))
                resolucion_obj = res_res.scalar_one_or_none()
                if not resolucion_obj:
                    err_msg = f"La resolución con código '{resolucion_raw}' no existe."
                    result.errors.append({"fila": fila_idx, "error": err_msg})
                    ws.cell(row=fila_idx, column=err_col_idx, value=err_msg)
                    await nested.rollback()
                    continue
                resolucion_id_db = resolucion_obj.id
            elif resolucion_id:
                resolucion_id_db = resolucion_id

            # Determinar categoría
            categoria = "TECNOLOGICO"
            tipo_upper = tipo_raw.upper()
            if "BIOMEDICO" in tipo_upper:
                categoria = "BIOMEDICO"
            elif "DOTACION" in tipo_upper or "INSUMO" in tipo_upper:
                err_msg = "Fila omitida: Dotación/Insumos deben importarse en el endpoint de cantidad."
                result.errors.append({"fila": fila_idx, "error": err_msg})
                ws.cell(row=fila_idx, column=err_col_idx, value=err_msg)
                await nested.rollback()
                continue

            # 1. Validar si artículo existe o crearlo
            q_art = select(Articulo).where(
                Articulo.categoria == categoria,
                Articulo.elemento == elemento_raw,
                Articulo.almacen_id == row_almacen_id
            )
            if marca_raw:
                q_art = q_art.where(Articulo.marca == marca_raw)
            if modelo_raw:
                q_art = q_art.where(Articulo.modelo == modelo_raw)
                
            res_art = await db.execute(q_art)
            articulo_obj = res_art.scalar_one_or_none()

            if not articulo_obj:
                articulo_obj = Articulo(
                    categoria=categoria,
                    tipo_elemento=tipo_raw.upper(),
                    elemento=elemento_raw,
                    marca=marca_raw,
                    modelo=modelo_raw,
                    requiere_serial=True,
                    stock_total=0,
                    stock_disponible=0,
                    almacen_id=row_almacen_id,
                    resolucion_id=resolucion_id_db
                )
                db.add(articulo_obj)
                await db.flush()
                result.created_articles += 1

            # 2. Validar duplicados por serial
            res_un = await db.execute(select(UnidadInventario).where(UnidadInventario.serial == serial_raw))
            if res_un.scalar_one_or_none():
                err_msg = f"El serial/IMEI '{serial_raw}' ya se encuentra registrado."
                result.skipped_duplicates += 1
                result.errors.append({"fila": fila_idx, "error": err_msg})
                ws.cell(row=fila_idx, column=err_col_idx, value=err_msg)
                await nested.rollback()
                continue

            # 3. Crear unidad
            nueva_unidad = UnidadInventario(
                articulo_id=articulo_obj.id,
                almacen_id=row_almacen_id,
                serial=serial_raw,
                imei2=imei2_raw,
                estado="DISPONIBLE",
                resolucion_id=resolucion_id_db
            )
            db.add(nueva_unidad)
            
            # Incrementar stocks del artículo
            articulo_obj.stock_total += 1
            articulo_obj.stock_disponible += 1
            
            await db.flush()
            result.created_units += 1
            ws.cell(row=fila_idx, column=err_col_idx, value="OK")
            await nested.commit()

        except Exception as e:
            await nested.rollback()
            err_msg = str(e)
            result.errors.append({"fila": fila_idx, "error": err_msg})
            ws.cell(row=fila_idx, column=err_col_idx, value=err_msg)

    if dry_run:
        await db.rollback()
    else:
        await db.commit()

    if result.errors:
        import uuid
        err_filename = f"reporte_errores_{uuid.uuid4().hex}.xlsx"
        err_filepath = os.path.join(settings.upload_dir, err_filename)
        wb.save(err_filepath)
        result.error_file = err_filename

    return result



@router.post("/import/cantidad", response_model=ImportInventarioResult)
async def importar_inventario_cantidad(
    file: UploadFile = File(...),
    almacen_id: Optional[int] = Query(None),
    resolucion_id: Optional[int] = Query(None),
    dry_run: bool = Query(False),
    db: AsyncSession = Depends(get_db)
):
    """Carga masiva de dotación e insumos (por cantidad de stock)."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "El archivo debe ser .xlsx")

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        ws = wb.active
        if ws is None:
            raise HTTPException(400, "El Excel no tiene hojas activas")
        err_col_idx = ws.max_column + 1
        ws.cell(row=1, column=err_col_idx, value="ERRORES")
    except Exception as e:
        raise HTTPException(400, f"Error al leer el Excel: {e}")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        raise HTTPException(400, "El archivo Excel está vacío")

    headers = [str(h) if h else "" for h in raw_headers]
    
    col_tipo = _find_col_idx(headers, "TIPO DE ELEMENTO")
    if col_tipo is None:
        col_tipo = _find_col_idx(headers, "TIPO")
    col_elemento = _find_col_idx(headers, "ELEMENTO")
    col_cantidad = _find_col_idx(headers, "CANTIDAD")
    col_almacen = _find_col_idx(headers, "ALMACEN_ID")
    if col_almacen is None:
        col_almacen = _find_col_idx(headers, "ID ALMACEN")
    if col_almacen is None:
        col_almacen = _find_col_idx(headers, "ALMACEN")
    if col_almacen is None:
        col_almacen = _find_col_idx(headers, "ID_ALMACEN")
    col_resolucion = _find_col_idx(headers, "RESOLUCION")
    if col_resolucion is None:
        col_resolucion = _find_col_idx(headers, "CODIGO RESOLUCION")
    if col_resolucion is None:
        col_resolucion = _find_col_idx(headers, "CODIGO_RESOLUCION")
    if col_resolucion is None:
        col_resolucion = _find_col_idx(headers, "RESOLUCION_CODIGO")
    if col_resolucion is None:
        col_resolucion = _find_col_idx(headers, "RESOLUCIÓN")

    if col_tipo is None or col_elemento is None or col_cantidad is None or col_almacen is None:
        raise HTTPException(
            400,
            "No se encontraron las columnas requeridas en el Excel: 'TIPO DE ELEMENTO', 'ELEMENTO', 'CANTIDAD' y 'ALMACEN_ID'."
        )

    result = ImportInventarioResult(dry_run=dry_run)
    for fila_idx, row in enumerate(rows_iter, start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        result.total += 1
        nested = await db.begin_nested()
        try:
            tipo_raw = _clean_str(row[col_tipo])
            elemento_raw = _clean_str(row[col_elemento])
            cantidad_raw = _parse_int(row[col_cantidad])
            row_almacen_id = _parse_int(row[col_almacen]) if col_almacen is not None else 0
            resolucion_raw = _clean_str(row[col_resolucion]) if col_resolucion is not None else None

            if not tipo_raw or not elemento_raw or cantidad_raw <= 0 or not row_almacen_id:
                err_msg = "Campos obligatorios vacíos o cantidad/ALMACEN_ID inválida"
                result.errors.append({"fila": fila_idx, "error": err_msg})
                ws.cell(row=fila_idx, column=err_col_idx, value=err_msg)
                await nested.rollback()
                continue

            # Validar si el almacén existe
            res_alm = await db.execute(select(Almacen).where(Almacen.id == row_almacen_id))
            almacen_obj = res_alm.scalar_one_or_none()
            if not almacen_obj:
                err_msg = f"El almacén con ID {row_almacen_id} no existe."
                result.errors.append({"fila": fila_idx, "error": err_msg})
                ws.cell(row=fila_idx, column=err_col_idx, value=err_msg)
                await nested.rollback()
                continue

            # Buscar la resolución por código si viene en la fila
            resolucion_id_db = None
            if resolucion_raw:
                res_res = await db.execute(select(Resolucion).where(Resolucion.codigo == resolucion_raw))
                resolucion_obj = res_res.scalar_one_or_none()
                if not resolucion_obj:
                    err_msg = f"La resolución con código '{resolucion_raw}' no existe."
                    result.errors.append({"fila": fila_idx, "error": err_msg})
                    ws.cell(row=fila_idx, column=err_col_idx, value=err_msg)
                    await nested.rollback()
                    continue
                resolucion_id_db = resolucion_obj.id
            elif resolucion_id:
                resolucion_id_db = resolucion_id

            # Validar que sea dotación
            tipo_upper = tipo_raw.upper()
            if "TECNOLOGICO" in tipo_upper or "BIOMEDICO" in tipo_upper:
                err_msg = "Fila omitida: Elementos tecnológicos/biomédicos deben importarse en el endpoint de serializados."
                result.errors.append({"fila": fila_idx, "error": err_msg})
                ws.cell(row=fila_idx, column=err_col_idx, value=err_msg)
                await nested.rollback()
                continue

            # Determine category
            categoria = "INSUMO" if "INSUMO" in tipo_upper else "DOTACION"

            # Buscar o crear artículo
            q_art = select(Articulo).where(
                Articulo.categoria == categoria,
                Articulo.elemento == elemento_raw,
                Articulo.almacen_id == row_almacen_id
            )
            res_art = await db.execute(q_art)
            articulo_obj = res_art.scalar_one_or_none()

            if not articulo_obj:
                articulo_obj = Articulo(
                    categoria=categoria,
                    tipo_elemento=tipo_raw.upper(),
                    elemento=elemento_raw,
                    requiere_serial=False,
                    stock_total=0,
                    stock_disponible=0,
                    almacen_id=row_almacen_id,
                    resolucion_id=resolucion_id_db
                )
                db.add(articulo_obj)
                await db.flush()
                result.created_articles += 1

            # Sumar cantidad al stock
            articulo_obj.stock_total += cantidad_raw
            articulo_obj.stock_disponible += cantidad_raw
            await db.flush()
            
            result.created_units += cantidad_raw
            ws.cell(row=fila_idx, column=err_col_idx, value="OK")
            await nested.commit()

        except Exception as e:
            await nested.rollback()
            err_msg = str(e)
            result.errors.append({"fila": fila_idx, "error": err_msg})
            ws.cell(row=fila_idx, column=err_col_idx, value=err_msg)

    if dry_run:
        await db.rollback()
    else:
        await db.commit()

    if result.errors:
        import uuid
        err_filename = f"reporte_errores_{uuid.uuid4().hex}.xlsx"
        err_filepath = os.path.join(settings.upload_dir, err_filename)
        wb.save(err_filepath)
        result.error_file = err_filename

    return result



# ─── ENDPOINTS DE ASIGNACIÓN (ENTREGA Y DEVOLUCIÓN) ───
@router.post("/entrega", response_model=List[ActaResponse])
async def registrar_entrega(
    req: EntregaRequest, 
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Registra la entrega de elementos a un contratista con contrato activo. Genera las actas .docx correspondientes."""
    # 1. Validar contrato
    res_cont = await db.execute(
        select(Contrato)
        .options(selectinload(Contrato.contratista_rel), selectinload(Contrato.resolucion))
        .where(Contrato.id == req.contrato_id)
    )
    contrato = res_cont.scalar_one_or_none()
    if not contrato:
        raise HTTPException(status_code=404, detail="Contrato no encontrado")
    if contrato.estado != "ACTIVO":
        raise HTTPException(status_code=400, detail="Solo se puede asignar inventario a contratos con estado ACTIVO")

    # Obtener detalles del contratista
    contratista_rel = contrato.contratista_rel
    contrato_data = {
        "numero_contrato": contrato.numero_contrato,
        "nombre_contratista": contratista_rel.nombre,
        "cedula": contratista_rel.identificacion,
        "perfil": contrato.perfil or "",
        "resolucion_codigo": contrato.resolucion.codigo if contrato.resolucion else ""
    }

    fecha_acta = req.fecha or date.today()
    recibido_por = req.recibido_por or contratista_rel.nombre
    recibido_entregado_por = current_user.nombre_completo or current_user.username

    # Agrupar items por categoría
    items_by_cat = {}
    
    # Procesar y validar cada item
    for item in req.items:
        res_art = await db.execute(select(Articulo).where(Articulo.id == item.articulo_id))
        articulo = res_art.scalar_one_or_none()
        if not articulo:
            raise HTTPException(status_code=404, detail=f"Artículo {item.articulo_id} no encontrado")

        cat_acta = "DOTACION" if articulo.categoria == "DOTACION_INSUMO" else articulo.categoria
        if cat_acta not in items_by_cat:
            items_by_cat[cat_acta] = []

        item_dict = {
            "articulo_id": articulo.id,
            "tipo_elemento": articulo.tipo_elemento,
            "elemento": articulo.elemento,
            "marca": articulo.marca or "N/A",
            "modelo": articulo.modelo or "N/A",
            "estado_declarado": item.estado_declarado or "Excelente",
            "observaciones": item.observaciones or ""
        }

        if articulo.requiere_serial:
            # Validar unidad
            if not item.unidad_id:
                raise HTTPException(status_code=400, detail=f"Debe especificar la unidad física para el artículo serializado '{articulo.elemento}'")
            res_un = await db.execute(select(UnidadInventario).where(UnidadInventario.id == item.unidad_id))
            unidad = res_un.scalar_one_or_none()
            if not unidad or unidad.articulo_id != articulo.id:
                raise HTTPException(status_code=404, detail="Unidad física no encontrada")
            if unidad.estado not in ESTADOS_DISPONIBLES:
                raise HTTPException(status_code=400, detail=f"La unidad con serial {unidad.serial} no está disponible (Estado actual: {unidad.estado})")

            # Asignar unidad
            unidad.estado = "ENTREGADO"
            unidad.contrato_actual_id = contrato.id
            unidad.fecha_ultima_entrega = fecha_acta
            unidad.observaciones = item.observaciones
            
            # Decrementar stock disponible
            articulo.stock_disponible = max(0, articulo.stock_disponible - 1)
            
            item_dict["unidad_id"] = unidad.id
            item_dict["serial"] = unidad.serial
            item_dict["imei2"] = unidad.imei2 or "N/A"
            
            items_by_cat[cat_acta].append((item_dict, unidad, articulo))
        else:
            # Validar stock
            qty = item.cantidad or 1
            if qty <= 0:
                raise HTTPException(status_code=400, detail="La cantidad debe ser mayor a 0")
            if articulo.stock_disponible < qty:
                raise HTTPException(status_code=400, detail=f"Stock insuficiente para '{articulo.elemento}'. Disponibles: {articulo.stock_disponible}, Solicitados: {qty}")

            # Decrementar stock disponible
            articulo.stock_disponible -= qty
            
            item_dict["cantidad"] = str(qty)
            items_by_cat[cat_acta].append((item_dict, None, articulo))

    # Guardar movimientos y generar actas
    actas_creadas = []
    os.makedirs(os.path.join(settings.upload_dir, "actas"), exist_ok=True)

    for cat, list_tuples in items_by_cat.items():
        # Crear Acta en BD
        nueva_acta = Acta(
            tipo="ENTREGA",
            categoria=cat,
            contrato_id=contrato.id,
            fecha=fecha_acta,
            recibido_entregado_por=recibido_entregado_por,
            usuario_id=current_user.id
        )
        db.add(nueva_acta)
        await db.flush()

        items_render = []
        for item_dict, unidad, articulo in list_tuples:
            # Crear Movimiento en BD
            movimiento = MovimientoInventario(
                tipo="ENTREGA",
                contrato_id=contrato.id,
                unidad_id=unidad.id if unidad else None,
                articulo_id=articulo.id,
                cantidad=int(item_dict.get("cantidad", 1)) if not unidad else 1,
                fecha=fecha_acta,
                estado_declarado=item_dict["estado_declarado"],
                observaciones=item_dict["observaciones"],
                recibido_por=recibido_por,
                acta_id=nueva_acta.id
            )
            db.add(movimiento)
            items_render.append(item_dict)

        # Generar archivo de acta .docx
        docx_bytes = generar_acta_inventario_docx(
            tipo="ENTREGA",
            categoria=cat,
            contrato_data=contrato_data,
            items=items_render,
            fecha_acta=fecha_acta
        )
        
        # Guardar archivo localmente
        filename = f"ACTA_ENTREGA_{cat}_{contrato.numero_contrato}_{nueva_acta.id}.docx".replace("/", "_")
        filepath = os.path.join(settings.upload_dir, "actas", filename)
        with open(filepath, "wb") as f:
            f.write(docx_bytes)

        # Actualizar ruta en Acta
        nueva_acta.archivo_generado = os.path.join("actas", filename)
        actas_creadas.append(nueva_acta)

    await db.commit()
    for a in actas_creadas:
        await db.refresh(a)
    return actas_creadas


@router.post("/devolucion", response_model=List[ActaResponse])
async def registrar_devolucion(
    req: DevolucionRequest, 
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Registra la devolución de elementos asignados a un contrato. Genera las actas .docx de devolución correspondientes."""
    res_cont = await db.execute(
        select(Contrato)
        .options(selectinload(Contrato.contratista_rel), selectinload(Contrato.resolucion))
        .where(Contrato.id == req.contrato_id)
    )
    contrato = res_cont.scalar_one_or_none()
    if not contrato:
        raise HTTPException(status_code=404, detail="Contrato no encontrado")

    contratista_rel = contrato.contratista_rel
    contrato_data = {
        "numero_contrato": contrato.numero_contrato,
        "nombre_contratista": contratista_rel.nombre,
        "cedula": contratista_rel.identificacion,
        "perfil": contrato.perfil or "",
        "resolucion_codigo": contrato.resolucion.codigo if contrato.resolucion else ""
    }

    fecha_acta = req.fecha or date.today()
    recibido_por = req.recibido_por or (contrato.supervisor or "Coordinación EBS")
    recibido_entregado_por = current_user.nombre_completo or current_user.username

    # Agrupar items por categoría
    items_by_cat = {}

    for item in req.items:
        if item.unidad_id:
            # Devolución de serializado
            res_un = await db.execute(
                select(UnidadInventario)
                .options(selectinload(UnidadInventario.articulo))
                .join(Articulo)
                .where(UnidadInventario.id == item.unidad_id)
            )
            unidad = res_un.scalar_one_or_none()
            if not unidad:
                raise HTTPException(status_code=404, detail=f"Unidad física {item.unidad_id} no encontrada")
            if unidad.contrato_actual_id != contrato.id:
                raise HTTPException(status_code=400, detail=f"La unidad {unidad.serial} no está asignada a este contrato")

            articulo = unidad.articulo
            cat_acta = "DOTACION" if articulo.categoria == "DOTACION_INSUMO" else articulo.categoria
            if cat_acta not in items_by_cat:
                items_by_cat[cat_acta] = []

            # Estado final de la unidad
            estado_final = "DISPONIBLE"
            est_dec = item.estado_declarado or "Excelente"
            est_dec_upper = est_dec.upper()
            if "DAÑO" in est_dec_upper or "REPARACION" in est_dec_upper or "FALLA" in est_dec_upper:
                estado_final = "EN_MANTENIMIENTO"
            elif "BAJA" in est_dec_upper or "PERDIDA" in est_dec_upper or "ROTO" in est_dec_upper:
                estado_final = "DE_BAJA"
            elif "REGULAR" in est_dec_upper:
                estado_final = "REGULAR"
            elif "BUEN" in est_dec_upper:
                estado_final = "BUEN_ESTADO"

            # Actualizar unidad
            unidad.estado = estado_final
            unidad.contrato_actual_id = None
            unidad.fecha_ultima_devolucion = fecha_acta
            
            # Si vuelve a estar disponible, se reintegra al stock disponible
            if estado_final in ESTADOS_DISPONIBLES:
                articulo.stock_disponible += 1
            elif estado_final == "DE_BAJA":
                # Se descuenta del stock total también
                articulo.stock_total = max(0, articulo.stock_total - 1)

            item_dict = {
                "articulo_id": articulo.id,
                "tipo_elemento": articulo.tipo_elemento,
                "elemento": articulo.elemento,
                "marca": articulo.marca or "N/A",
                "modelo": articulo.modelo or "N/A",
                "unidad_id": unidad.id,
                "serial": unidad.serial,
                "imei2": unidad.imei2 or "N/A",
                "estado_declarado": est_dec,
                "observaciones": item.observaciones or ""
            }
            items_by_cat[cat_acta].append((item_dict, unidad, articulo, True))

        elif item.articulo_id:
            # Devolución de dotación/insumo
            res_art = await db.execute(select(Articulo).where(Articulo.id == item.articulo_id))
            articulo = res_art.scalar_one_or_none()
            if not articulo or articulo.requiere_serial:
                raise HTTPException(status_code=400, detail="Artículo inválido para devolución sin serial")

            cat_acta = "DOTACION" if articulo.categoria == "DOTACION_INSUMO" else articulo.categoria
            if cat_acta not in items_by_cat:
                items_by_cat[cat_acta] = []

            qty = item.cantidad or 1
            reutilizable = item.reutilizable if item.reutilizable is not None else True
            
            if reutilizable:
                # Reingresa a stock
                articulo.stock_disponible += qty
            else:
                # Se da de baja (sale del stock total)
                articulo.stock_total = max(0, articulo.stock_total - qty)

            item_dict = {
                "articulo_id": articulo.id,
                "tipo_elemento": articulo.tipo_elemento,
                "elemento": articulo.elemento,
                "cantidad": str(qty),
                "estado_declarado": item.estado_declarado or "Excelente",
                "observaciones": f"{item.observaciones or ''} (Reutilizable: {'SI' if reutilizable else 'NO'})"
            }
            items_by_cat[cat_acta].append((item_dict, None, articulo, reutilizable))
        else:
            raise HTTPException(status_code=400, detail="Debe especificar unidad_id (serializado) o articulo_id (cantidad) para la devolución")

    actas_creadas = []
    os.makedirs(os.path.join(settings.upload_dir, "actas"), exist_ok=True)

    for cat, list_tuples in items_by_cat.items():
        # Crear Acta en BD
        nueva_acta = Acta(
            tipo="DEVOLUCION",
            categoria=cat,
            contrato_id=contrato.id,
            fecha=fecha_acta,
            recibido_entregado_por=recibido_entregado_por,
            usuario_id=current_user.id
        )
        db.add(nueva_acta)
        await db.flush()

        items_render = []
        for item_dict, unidad, articulo, reutilizable in list_tuples:
            # Crear Movimiento en BD
            movimiento = MovimientoInventario(
                tipo="DEVOLUCION",
                contrato_id=contrato.id,
                unidad_id=unidad.id if unidad else None,
                articulo_id=articulo.id,
                cantidad=int(item_dict.get("cantidad", 1)) if not unidad else 1,
                fecha=fecha_acta,
                estado_declarado=item_dict["estado_declarado"],
                observaciones=item_dict["observaciones"],
                recibido_por=recibido_por,
                acta_id=nueva_acta.id
            )
            db.add(movimiento)
            items_render.append(item_dict)

        # Generar archivo de acta .docx
        docx_bytes = generar_acta_inventario_docx(
            tipo="DEVOLUCION",
            categoria=cat,
            contrato_data=contrato_data,
            items=items_render,
            fecha_acta=fecha_acta
        )
        
        # Guardar archivo localmente
        filename = f"ACTA_DEVOLUCION_{cat}_{contrato.numero_contrato}_{nueva_acta.id}.docx".replace("/", "_")
        filepath = os.path.join(settings.upload_dir, "actas", filename)
        with open(filepath, "wb") as f:
            f.write(docx_bytes)

        # Actualizar ruta en Acta
        nueva_acta.archivo_generado = os.path.join("actas", filename)
        actas_creadas.append(nueva_acta)

    await db.commit()
    for a in actas_creadas:
        await db.refresh(a)
    return actas_creadas


# ─── ENDPOINTS DE DETALLE Y CONSULTA ───
@router.get("/contrato/{contrato_id}")
async def detalle_inventario_contrato(contrato_id: int, db: AsyncSession = Depends(get_db)):
    """Obtiene el inventario actual asignado y el histórico de movimientos/actas de un contrato."""
    res_cont = await db.execute(
        select(Contrato)
        .options(selectinload(Contrato.contratista_rel))
        .where(Contrato.id == contrato_id)
    )
    contrato = res_cont.scalar_one_or_none()
    if not contrato:
        raise HTTPException(status_code=404, detail="Contrato no encontrado")

    # 1. Unidades serializadas activas
    res_unidades = await db.execute(
        select(UnidadInventario)
        .options(selectinload(UnidadInventario.articulo))
        .join(Articulo)
        .where(UnidadInventario.contrato_actual_id == contrato.id)
    )
    unidades = res_unidades.scalars().all()

    # 2. Dotación acumulada
    # Suma entregado - devuelto por artículo para este contrato
    res_movs = await db.execute(
        select(MovimientoInventario)
        .options(
            selectinload(MovimientoInventario.articulo),
            selectinload(MovimientoInventario.unidad)
        )
        .join(Articulo)
        .where(MovimientoInventario.contrato_id == contrato.id)
        .order_by(MovimientoInventario.fecha.desc())
    )
    todos_movimientos = res_movs.scalars().all()

    # Calcular stock neto asignado de dotación
    dotacion_asignada = {}
    for m in todos_movimientos:
        if m.articulo and not m.articulo.requiere_serial:
            art_id = m.articulo.id
            if art_id not in dotacion_asignada:
                dotacion_asignada[art_id] = {
                    "articulo_id": art_id,
                    "elemento": m.articulo.elemento,
                    "tipo_elemento": m.articulo.tipo_elemento,
                    "categoria": m.articulo.categoria,
                    "cantidad_neta": 0
                }
            if m.tipo == "ENTREGA":
                dotacion_asignada[art_id]["cantidad_neta"] += m.cantidad
            else:
                dotacion_asignada[art_id]["cantidad_neta"] -= m.cantidad

    # Filtrar solo las dotaciones con cantidad_neta > 0
    list_dotaciones = [v for v in dotacion_asignada.values() if v["cantidad_neta"] > 0]

    # 3. Actas
    res_actas = await db.execute(
        select(Acta).where(Acta.contrato_id == contrato.id).order_by(Acta.fecha.desc())
    )
    actas = res_actas.scalars().all()

    return {
        "contrato_id": contrato.id,
        "numero_contrato": contrato.numero_contrato,
        "contratista": contrato.contratista_rel.nombre if contrato.contratista_rel else "",
        "equipos_asignados": [
            {
                "id": u.id,
                "serial": u.serial,
                "imei2": u.imei2,
                "estado": u.estado,
                "elemento": u.articulo.elemento,
                "tipo_elemento": u.articulo.tipo_elemento,
                "categoria": u.articulo.categoria,
                "marca": u.articulo.marca,
                "modelo": u.articulo.modelo,
                "fecha_ultima_entrega": u.fecha_ultima_entrega
            } for u in unidades
        ],
        "dotaciones_asignadas": list_dotaciones,
        "historial_movimientos": [
            {
                "id": m.id,
                "tipo": m.tipo,
                "elemento": m.articulo.elemento if m.articulo else "N/A",
                "serial": m.unidad.serial if m.unidad else None,
                "cantidad": m.cantidad,
                "fecha": m.fecha,
                "estado_declarado": m.estado_declarado,
                "observaciones": m.observaciones,
                "recibido_por": m.recibido_por,
                "acta_id": m.acta_id
            } for m in todos_movimientos
        ],
        "actas": [
            {
                "id": a.id,
                "tipo": a.tipo,
                "categoria": a.categoria,
                "fecha": a.fecha,
                "archivo_generado": a.archivo_generado,
                "recibido_entregado_por": a.recibido_entregado_por
            } for a in actas
        ]
    }


@router.get("/actas/contratista/{contratista_id}", response_model=List[ActaListResponse])
async def obtener_actas_contratista(contratista_id: int, db: AsyncSession = Depends(get_db)):
    """Obtiene el listado de actas de inventario asociadas a los contratos de un contratista."""
    stmt = (
        select(Acta)
        .join(Contrato)
        .options(selectinload(Acta.contrato))
        .where(Contrato.contratista_id == contratista_id)
        .order_by(Acta.fecha.desc(), Acta.created_at.desc())
    )
    result = await db.execute(stmt)
    actas = result.scalars().all()

    # Map the contract number dynamically
    for a in actas:
        a.numero_contrato = a.contrato.numero_contrato if a.contrato else None

    return actas


@router.get("/actas/{acta_id}/pdf")
async def descargar_acta_pdf(acta_id: int, descargar: bool = Query(False), db: AsyncSession = Depends(get_db)):
    """Genera y descarga el documento de un acta en formato PDF."""
    from app.services.pdf_generator import generar_acta_pdf
    from app.models.contratista import Contratista
    from app.models.resolucion import Resolucion

    # Cargar Acta con relaciones
    stmt = (
        select(Acta)
        .options(
            selectinload(Acta.contrato).selectinload(Contrato.contratista_rel),
            selectinload(Acta.contrato).selectinload(Contrato.resolucion),
            selectinload(Acta.movimientos).selectinload(MovimientoInventario.unidad).selectinload(UnidadInventario.articulo),
            selectinload(Acta.movimientos).selectinload(MovimientoInventario.articulo)
        )
        .where(Acta.id == acta_id)
    )
    res_acta = await db.execute(stmt)
    acta = res_acta.scalar_one_or_none()
    if not acta:
        raise HTTPException(status_code=404, detail="Acta no encontrada")

    contrato = acta.contrato
    if not contrato:
        raise HTTPException(status_code=404, detail="Contrato asociado no encontrado")

    contratista = contrato.contratista_rel
    if not contratista:
        raise HTTPException(status_code=404, detail="Contratista asociado no encontrado")

    # Mapear items
    items_list = []
    for mov in acta.movimientos:
        articulo_nombre = ""
        marca = ""
        modelo = ""
        serial = ""
        imei2 = ""

        if mov.unidad:
            articulo_nombre = mov.unidad.articulo.elemento if mov.unidad.articulo else ""
            marca = mov.unidad.articulo.marca if mov.unidad.articulo else ""
            modelo = mov.unidad.articulo.modelo if mov.unidad.articulo else ""
            serial = mov.unidad.serial or ""
            imei2 = mov.unidad.imei2 or "N/A"
        elif mov.articulo:
            articulo_nombre = mov.articulo.elemento
            marca = mov.articulo.marca or ""
            modelo = mov.articulo.modelo or ""

        items_list.append({
            "elemento": articulo_nombre,
            "marca": marca,
            "modelo": modelo,
            "serial": serial,
            "imei2": imei2,
            "cantidad": mov.cantidad,
            "estado_declarado": mov.estado_declarado or "Excelente",
            "observaciones": mov.observaciones or ""
        })

    contratista_dict = {
        "nombre": contratista.nombre,
        "identificacion": contratista.identificacion
    }

    resolucion_codigo = contrato.resolucion.codigo if contrato.resolucion else ""

    contrato_dict = {
        "numero_contrato": contrato.numero_contrato,
        "perfil": contrato.perfil,
        "resolucion_codigo": resolucion_codigo
    }

    logo_b64 = ""

    pdf_bytes = generar_acta_pdf(
        acta_tipo=acta.tipo,
        categoria=acta.categoria,
        contratista=contratista_dict,
        contrato=contrato_dict,
        items=items_list,
        fecha_acta=acta.fecha,
        recibido_entregado_por=acta.recibido_entregado_por,
        logo_b64=logo_b64
    )

    filename = f"ACTA_{acta.tipo}_{acta.categoria}_{contrato.numero_contrato}_{acta.id}.pdf".replace("/", "_")
    disposition = "attachment" if descargar else "inline"
    return Response(
        pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"{disposition}; filename={filename}"}
    )


@router.get("/actas/{acta_id}")
async def obtener_detalle_acta(acta_id: int, db: AsyncSession = Depends(get_db)):
    """Obtiene los detalles y movimientos de un acta específica."""
    stmt = (
        select(Acta)
        .options(
            selectinload(Acta.contrato),
            selectinload(Acta.movimientos).selectinload(MovimientoInventario.unidad).selectinload(UnidadInventario.articulo),
            selectinload(Acta.movimientos).selectinload(MovimientoInventario.articulo)
        )
        .where(Acta.id == acta_id)
    )
    result = await db.execute(stmt)
    acta = result.scalar_one_or_none()
    if not acta:
        raise HTTPException(status_code=404, detail="Acta no encontrada")

    # Mapear movimientos
    movs_list = []
    for mov in acta.movimientos:
        elemento = ""
        marca = ""
        modelo = ""
        serial = ""
        imei2 = ""

        if mov.unidad:
            elemento = mov.unidad.articulo.elemento if mov.unidad.articulo else ""
            marca = mov.unidad.articulo.marca if mov.unidad.articulo else ""
            modelo = mov.unidad.articulo.modelo if mov.unidad.articulo else ""
            serial = mov.unidad.serial or ""
            imei2 = mov.unidad.imei2 or ""
        elif mov.articulo:
            elemento = mov.articulo.elemento
            marca = mov.articulo.marca or ""
            modelo = mov.articulo.modelo or ""

        movs_list.append({
            "id": mov.id,
            "elemento": elemento,
            "marca": marca,
            "modelo": modelo,
            "serial": serial,
            "imei2": imei2,
            "cantidad": mov.cantidad,
            "estado_declarado": mov.estado_declarado,
            "observaciones": mov.observaciones
        })

    return {
        "id": acta.id,
        "tipo": acta.tipo,
        "categoria": acta.categoria,
        "fecha": acta.fecha,
        "recibido_entregado_por": acta.recibido_entregado_por,
        "numero_contrato": acta.contrato.numero_contrato if acta.contrato else None,
        "movimientos": movs_list
    }


@router.get("/actas/{acta_id}/download")
async def descargar_acta(acta_id: int, db: AsyncSession = Depends(get_db)):
    """Descarga el documento de un acta en formato .docx."""
    res_acta = await db.execute(select(Acta).where(Acta.id == acta_id))
    acta = res_acta.scalar_one_or_none()
    if not acta or not acta.archivo_generado:
        raise HTTPException(status_code=404, detail="Acta o archivo no encontrado")

    filepath = os.path.join(settings.upload_dir, acta.archivo_generado)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail=f"El archivo físico no existe en la ruta {filepath}")

    filename = os.path.basename(filepath)
    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename=filename
    )


@router.get("/import/errores/{filename}/download")
async def descargar_reporte_errores(filename: str):
    """Descarga el reporte de errores generado en una importación fallida."""
    safe_filename = os.path.basename(filename)
    filepath = os.path.join(settings.upload_dir, safe_filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Archivo de reporte de errores no encontrado")

    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=safe_filename
    )


@router.get("/unidades/{unidad_id}/historial", response_model=List[MovimientoResponse])
async def historial_unidad(unidad_id: int, db: AsyncSession = Depends(get_db)):
    """Obtiene el historial de movimientos de una unidad específica."""
    res = await db.execute(
        select(MovimientoInventario)
        .where(MovimientoInventario.unidad_id == unidad_id)
        .order_by(MovimientoInventario.fecha.desc())
    )
    return res.scalars().all()


# ─── DASHBOARD DE DISPONIBILIDAD ───
@router.get("/dashboard", response_model=DashboardResponse)
async def dashboard_disponibilidad(db: AsyncSession = Depends(get_db)):
    """Métricas globales de inventario y estado actual por artículo."""
    # 1. Agrupar por categoría
    # Para tecnológico y biomédico sumamos las unidades de UnidadInventario
    res_unidades = await db.execute(
        select(
            Articulo.categoria,
            UnidadInventario.estado,
            func.count(UnidadInventario.id)
        )
        .join(UnidadInventario)
        .group_by(Articulo.categoria, UnidadInventario.estado)
    )
    counts = res_unidades.all()

    resumen = {
        "TECNOLOGICO": CategoriaResumen(),
        "BIOMEDICO": CategoriaResumen(),
        "DOTACION_INSUMO": CategoriaResumen(),
        "DOTACION": CategoriaResumen(),
        "INSUMO": CategoriaResumen()
    }

    for cat, estado, count in counts:
        if cat in resumen:
            r = resumen[cat]
            r.total += count
            if estado in ESTADOS_DISPONIBLES:
                r.disponibles += count
            elif estado == "ENTREGADO":
                r.entregados += count
            elif estado in {"EN_MANTENIMIENTO", "DANADO"}:
                r.mantenimiento += count
            elif estado == "DE_BAJA":
                r.baja += count

    # Para dotación/insumos sumamos directamente el stock_total y stock_disponible
    res_dotacion = await db.execute(
        select(
            Articulo.categoria,
            func.sum(Articulo.stock_total),
            func.sum(Articulo.stock_disponible)
        )
        .where(Articulo.categoria.in_(["DOTACION_INSUMO", "DOTACION", "INSUMO"]))
        .group_by(Articulo.categoria)
    )
    for cat, total_d, disp_d in res_dotacion.all():
        if cat in resumen:
            r_d = resumen[cat]
            r_d.total = int(total_d or 0)
            r_d.disponibles = int(disp_d or 0)
            r_d.entregados = max(0, r_d.total - r_d.disponibles)

    # Combinamos legacy DOTACION_INSUMO en DOTACION
    resumen["DOTACION"].total += resumen["DOTACION_INSUMO"].total
    resumen["DOTACION"].disponibles += resumen["DOTACION_INSUMO"].disponibles
    resumen["DOTACION"].entregados += resumen["DOTACION_INSUMO"].entregados

    # 2. Detalle por artículo
    res_articulos = await db.execute(select(Articulo).order_by(Articulo.elemento))
    articulos = res_articulos.scalars().all()

    lista_articulos_info = []
    for art in articulos:
        if art.requiere_serial:
            # Calcular estados desde unidades
            res_un_art = await db.execute(
                select(UnidadInventario.estado, func.count(UnidadInventario.id))
                .where(UnidadInventario.articulo_id == art.id)
                .group_by(UnidadInventario.estado)
            )
            states = dict(res_un_art.all())
            
            registrados = sum(states.values())
            entregados = states.get("ENTREGADO", 0)
            disponibles = states.get("DISPONIBLE", 0) + states.get("BUEN_ESTADO", 0) + states.get("REGULAR", 0)
            mantenimiento = states.get("EN_MANTENIMIENTO", 0) + states.get("DANADO", 0)
            baja = states.get("DE_BAJA", 0)
        else:
            registrados = art.stock_total
            disponibles = art.stock_disponible
            entregados = registrados - disponibles
            mantenimiento = 0
            baja = 0

        lista_articulos_info.append(
            ArticuloDashboardInfo(
                id=art.id,
                categoria=art.categoria,
                tipo_elemento=art.tipo_elemento,
                elemento=art.elemento,
                marca=art.marca,
                modelo=art.modelo,
                registrados=registrados,
                entregados=entregados,
                disponibles=disponibles,
                mantenimiento=mantenimiento,
                baja=baja,
                resolucion_id=art.resolucion_id
            )
        )

    return DashboardResponse(
        resumen_tecnologico=resumen["TECNOLOGICO"],
        resumen_biomedico=resumen["BIOMEDICO"],
        resumen_dotacion=resumen["DOTACION"],
        resumen_insumo=resumen["INSUMO"],
        articulos=lista_articulos_info
    )


@router.get("/unidades/export")
async def exportar_unidades_excel_endpoint(
    articulo_id: Optional[int] = None,
    almacen_id: Optional[int] = None,
    estado: Optional[str] = None,
    categoria: Optional[str] = None,
    elemento: Optional[str] = None,
    asignado_a_contrato: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """Exporta las unidades físicas a Excel aplicando filtros."""
    q = select(UnidadInventario).options(
        selectinload(UnidadInventario.articulo),
        selectinload(UnidadInventario.contrato_actual).selectinload(Contrato.contratista_rel)
    ).join(Articulo)
    if articulo_id:
        q = q.where(UnidadInventario.articulo_id == articulo_id)
    if almacen_id:
        q = q.where(UnidadInventario.almacen_id == almacen_id)
    if estado:
        q = q.where(UnidadInventario.estado == estado)
    if categoria:
        q = q.where(Articulo.categoria == categoria)
    if elemento:
        q = q.where(Articulo.elemento.ilike(f"%{elemento}%"))
    if asignado_a_contrato:
        if asignado_a_contrato == "SI":
            q = q.where(UnidadInventario.contrato_actual_id.isnot(None))
        elif asignado_a_contrato == "NO":
            q = q.where(UnidadInventario.contrato_actual_id.is_(None))
    if search:
        search_filter = f"%{search}%"
        q = q.where(
            (UnidadInventario.serial.ilike(search_filter)) |
            (Articulo.elemento.ilike(search_filter)) |
            (UnidadInventario.estado.ilike(search_filter))
        )
        
    res = await db.execute(q.order_by(UnidadInventario.serial))
    unidades = res.scalars().all()
    
    rows = []
    for u in unidades:
        rows.append({
            "categoria": u.articulo.categoria if u.articulo else "",
            "elemento": u.articulo.elemento if u.articulo else "",
            "serial": u.serial,
            "imei2": u.imei2,
            "estado": u.estado,
            "contrato_id": u.contrato_actual_id,
            "contratista_nombre": u.contratista_nombre
        })
        
    from app.services.excel_service import exportar_unidades_excel
    from fastapi.responses import Response
    excel_bytes = exportar_unidades_excel(rows)
    
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventario_unidades.xlsx"},
    )


@router.get("/dashboard/export")
async def exportar_dashboard_excel_endpoint(
    categoria: Optional[str] = None,
    almacen_id: Optional[int] = None,
    elemento: Optional[str] = None,
    marca_modelo: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """Exporta el resumen de disponibilidad general a Excel aplicando filtros."""
    q = select(Articulo)
    if categoria:
        q = q.where(Articulo.categoria == categoria)
    if almacen_id:
        q = q.where(Articulo.almacen_id == almacen_id)
    if elemento:
        q = q.where(Articulo.elemento.ilike(f"%{elemento}%"))
    if marca_modelo:
        term = f"%{marca_modelo}%"
        q = q.where((Articulo.marca.ilike(term)) | (Articulo.modelo.ilike(term)))
    if search:
        s_term = f"%{search}%"
        q = q.where(
            (Articulo.elemento.ilike(s_term)) |
            (Articulo.tipo_elemento.ilike(s_term)) |
            (Articulo.marca.ilike(s_term))
        )
        
    res_articulos = await db.execute(q.order_by(Articulo.elemento))
    articulos = res_articulos.scalars().all()
    
    rows = []
    for art in articulos:
        if art.requiere_serial:
            # Calcular estados desde unidades
            unidades_stmt = select(UnidadInventario.estado, func.count(UnidadInventario.id)).where(UnidadInventario.articulo_id == art.id)
            if almacen_id:
                unidades_stmt = unidades_stmt.where(UnidadInventario.almacen_id == almacen_id)
            res_un_art = await db.execute(unidades_stmt.group_by(UnidadInventario.estado))
            states = dict(res_un_art.all())
            
            registrados = sum(states.values())
            entregados = states.get("ENTREGADO", 0)
            disponibles = states.get("DISPONIBLE", 0) + states.get("BUEN_ESTADO", 0) + states.get("REGULAR", 0)
            mantenimiento = states.get("EN_MANTENIMIENTO", 0) + states.get("DANADO", 0)
            baja = states.get("DE_BAJA", 0)
        else:
            registrados = art.stock_total
            disponibles = art.stock_disponible
            entregados = registrados - disponibles
            mantenimiento = 0
            baja = 0
            
        rows.append({
            "categoria": art.categoria,
            "elemento": art.elemento,
            "tipo_elemento": art.tipo_elemento,
            "marca": art.marca,
            "modelo": art.modelo,
            "registrados": registrados,
            "disponibles": disponibles,
            "entregados": entregados,
            "mantenimiento": mantenimiento,
            "baja": baja
        })
        
    from app.services.excel_service import exportar_disponibilidad_excel
    from fastapi.responses import Response
    excel_bytes = exportar_disponibilidad_excel(rows)
    
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=disponibilidad_inventario.xlsx"},
    )

