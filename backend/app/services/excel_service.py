"""Exportación a Excel de contratos por resolución."""

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import text


def exportar_resolucion_excel(rows: list[dict]) -> bytes:
    """Genera un archivo Excel con los contratos de una resolución.

    Args:
        rows: Lista de dicts con datos de contratos

    Returns:
        bytes del archivo .xlsx
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contratos"

    VERDE = "1A7A4A"
    BORDE = Side(style="thin", color="CCCCCC")
    borde = Border(left=BORDE, right=BORDE, top=BORDE, bottom=BORDE)

    # Título
    ws.merge_cells("A1:R1")
    t = ws["A1"]
    t.value = f"CONTRATOS GENERADOS"
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=VERDE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Encabezados
    headers = [
        "No.", "No. Contrato", "Beneficiario", "Cédula", "Perfil", "Objeto",
        "Valor Total", "Valor Transporte", "CDP", "Fecha CDP", "Rubro",
        "Fecha Inicio", "Fecha Fin", "Estado", "Supervisor", "Unidad Atención",
        "Cuotas", "Cuotas Pagadas",
    ]
    header_fill = PatternFill("solid", fgColor="E8F5EE")
    header_font = Font(name="Arial", bold=True, size=10, color="1A7A4A")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
    ws.row_dimensions[2].height = 30

    # Datos
    for i, row in enumerate(rows, 3):
        data = [
            i - 2,
            row.get("numero_contrato", ""),
            row.get("beneficiario", ""),
            row.get("cedula_contratista", ""),
            row.get("perfil", ""),
            row.get("objeto", ""),
            row.get("monto_total", 0),
            row.get("monto_transporte", 0),
            row.get("no_cdp", ""),
            row.get("fecha_cdp", ""),
            row.get("rubro", ""),
            row.get("fecha_inicio", ""),
            row.get("fecha_fin", ""),
            row.get("estado", ""),
            row.get("supervisor", ""),
            row.get("unidad_atencion", ""),
            row.get("cuotas", ""),
            row.get("cuotas_pagadas", 0),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = borde
            if col in (7, 8) and isinstance(value, (int, float)) and value > 0:
                cell.number_format = '$#,##0'

    # Ancho de columnas
    widths = [5, 15, 30, 15, 20, 40, 15, 15, 15, 12, 15, 12, 12, 12, 25, 20, 10, 12]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def exportar_unidades_excel(rows: list[dict]) -> bytes:
    """Genera un archivo Excel con las unidades físicas (seriales).

    Args:
        rows: Lista de dicts con datos de unidades físicas

    Returns:
        bytes del archivo .xlsx
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unidades Fisicas"

    VERDE = "1A7A4A"
    BORDE = Side(style="thin", color="CCCCCC")
    borde = Border(left=BORDE, right=BORDE, top=BORDE, bottom=BORDE)

    # Título
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "REPORTE DE UNIDADES FÍSICAS (SERIALES)"
    t.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=VERDE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Encabezados
    headers = [
        "Categoría", "Elemento", "Serial / IMEI 1", "IMEI 2", "Estado", "Asignado a Contrato"
    ]
    header_fill = PatternFill("solid", fgColor="E8F5EE")
    header_font = Font(name="Arial", bold=True, size=10, color="1A7A4A")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
    ws.row_dimensions[2].height = 25

    # Datos
    for i, row in enumerate(rows, 3):
        contrato_val = row.get("contratista_nombre") or "—"
        data = [
            row.get("categoria", ""),
            row.get("elemento", ""),
            row.get("serial", ""),
            row.get("imei2", "") or "—",
            row.get("estado", ""),
            contrato_val
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = borde

    # Ancho de columnas
    widths = [15, 30, 20, 20, 15, 25]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def exportar_disponibilidad_excel(rows: list[dict]) -> bytes:
    """Genera un archivo Excel con la disponibilidad general por elemento.

    Args:
        rows: Lista de dicts con datos de artículos y stocks

    Returns:
        bytes del archivo .xlsx
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Disponibilidad General"

    VERDE = "1A7A4A"
    BORDE = Side(style="thin", color="CCCCCC")
    borde = Border(left=BORDE, right=BORDE, top=BORDE, bottom=BORDE)

    # Título
    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "ESTADO DE DISPONIBILIDAD DE INVENTARIO POR ELEMENTO"
    t.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=VERDE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Encabezados
    headers = [
        "Categoría", "Elemento", "Tipo Elemento", "Marca", "Modelo", 
        "Registrados", "Disponibles", "Entregados", "En Mantenimiento", "De Baja"
    ]
    header_fill = PatternFill("solid", fgColor="E8F5EE")
    header_font = Font(name="Arial", bold=True, size=10, color="1A7A4A")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
    ws.row_dimensions[2].height = 25

    # Datos
    for i, row in enumerate(rows, 3):
        data = [
            row.get("categoria", ""),
            row.get("elemento", ""),
            row.get("tipo_elemento", ""),
            row.get("marca", "") or "—",
            row.get("modelo", "") or "—",
            row.get("registrados", 0),
            row.get("disponibles", 0),
            row.get("entregados", 0),
            row.get("mantenimiento", 0),
            row.get("baja", 0)
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = borde
            if col > 5:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ancho de columnas
    widths = [15, 30, 20, 15, 15, 12, 12, 12, 16, 12]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
